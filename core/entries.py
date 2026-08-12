"""Entry list: import from Excel, overlay of jury edits, validation, export.

The file the federation sends stays the source of truth and is never written
to. Import produces a read-only snapshot; every edit made in the app is
recorded as a `Patch` in a separate overlay file and re-applied on top of each
new import, so a fresh export from the federation never silently discards jury
decisions. Patches are keyed by **UCI ID**, the one code that does not change:
a dorsale or a regione corrected in the app survives any number of reloads.

`import_entries` reads either shape, told apart by what the workbook holds:

* **the ksport export** (`Iscritti_NNNNNN_KSPORT.xlsx`) - one flat sheet, one
  row per rider, no event columns at all. Which specialità a rider contests is
  not in this file: the programme says which columns a categoria has, and the
  jury ticks them at the verifica (they live in the overlay).
* **the master workbook** (`Iscritti_26_generale.xlsx`) - a `KSPORT` sheet plus one
  printable sheet per category, whose event columns hold `X` (starter), `R`
  (reserve) or a pairing letter/number, and a `PROVINCE` sigla -> region lookup.

Column headings are never hard-coded: `entries.columns` and `entries.ksport` in
programme.yaml map the file's Italian headings to the field names used here,
matched by name rather than by position, so a shifted column does not silently
import garbage and a differently-worded export is a config change.

`entries.check_in` is the third mapping and the odd one out: the two columns of
the licence check (`Verificato`, `NP`) are not in the federation's layout at
all - the giuria adds them to the workbook itself. Wherever they are, on the
category sheets or on the KSPORT one, they are read on import and written back
by `write_back`; where they are not, the check lives in the overlay as before.
"""

from __future__ import annotations

import hashlib
from dataclasses import asdict, dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable

from .config import Competition, EVENT_ENTRY_LIST, EntrySheet
from .checks import ERROR, WARN, Issue
from .i18n import fix_accents, label, msg
from .models import (EntryList, EventEntry, FLAG_RESERVE, LEGACY_KEYS,
                     Pair, Rider, Team, pairing_letter, pairing_number)


# ── helpers ─────────────────────────────────────────────────────────────────

def _s(v: Any) -> str:
    """Cell -> trimmed string (''for None / NaN / Excel errors)."""
    if v is None:
        return ""
    if isinstance(v, float) and v != v:
        return ""
    s = str(v).strip()
    if s.startswith("#") and s.endswith(("!", "A", "?", "0")):  # N/A, #VALUE!
        return ""
    # `######` is not a value: it is what Excel caches for a cell whose column
    # is too narrow to show it. The 2026 export writes it in every certificate
    # date, and read as text it would raise a warning per rider.
    if set(s) == {"#"}:
        return ""
    return fix_accents(s)


def _int(v: Any) -> int | None:
    s = _s(v)
    if not s:
        return None
    try:
        return int(float(s))
    except ValueError:
        return None


def _uci(v: Any) -> str:
    """UCI ID as a plain digit string (regional forms store it as a number)."""
    s = _s(v)
    if s.endswith(".0"):
        s = s[:-2]
    return "".join(ch for ch in s if ch.isdigit())


#: What a ticked cell is written as. The workbook says SI / NO in its own
#: columns (`Riserva`), so the licence check speaks the same language: a
#: checkbox written as TRUE would read as English in an Italian file.
YES = "SI"

#: Everything read back as ticked: the app writes SI, the giuria types whatever
#: is at hand - an X, a flag typed in Excel as a boolean.
YES_VALUES = ("SI", "S", "YES", "X", "TRUE", "VERO", "1")


def _yes(v: Any) -> bool:
    """Is this cell a tick? (blank, NO and anything else are not)."""
    return _s(v).upper() in YES_VALUES


def _date(v: Any) -> str:
    """Cell -> ISO date. The export writes a timestamp; a sheet reads a date."""
    if isinstance(v, datetime):
        return v.date().isoformat()
    s = _s(v)
    head = s.split(" ")[0]
    try:
        return datetime.fromisoformat(head).date().isoformat()
    except ValueError:
        return s


#: Invisible characters an entry form picks up from a copy-paste: the cell looks
#: empty and is not. Left in, they make a region that is blank to the eye but
#: passes every "is it filled in?" test - and quietly becomes a squadra of its
#: own. The 2026 file has them on the two riders licensed abroad.
INVISIBLE = "​‌‍‎‏‪‫‬‭‮﻿\xa0"


def norm_region(s: str) -> str:
    # the export writes "?" where the federal form has no regione (a rider
    # licensed abroad): that is a blank, and the validation says so
    out = " ".join(_s(s).translate({ord(c): " " for c in INVISIBLE}).upper().split())
    return "" if out in ("?", "-") else out


def rider_key(uci_id: str, cat: str, bib: int | None) -> str:
    return uci_id if uci_id else f"{cat}-{bib}"


def file_hash(path: str | Path) -> str:
    h = hashlib.sha256()
    with open(path, "rb") as fh:
        for chunk in iter(lambda: fh.read(65536), b""):
            h.update(chunk)
    return h.hexdigest()[:16]


# Long-form category names used by some regional forms.
CAT_ALIASES = {
    "ALLIEVO": "AL", "ALLIEVI": "AL", "UOMINI ALLIEVI": "AL",
    "ALLIEVO F.": "DA", "ALLIEVA": "DA", "DONNE ALLIEVE": "DA", "ALLIEVE": "DA",
    "ESORDIENTE": "ES", "ESORDIENTI": "ES", "UOMINI ESORDIENTI": "ES",
    "ESORDIENTE F.": "ED", "ESORDIENTE F": "ED", "DONNE ESORDIENTI": "ED",
}


def norm_cat(s: str) -> str:
    s = _s(s).upper().rstrip(". ")
    if s in ("AL", "DA", "ES", "ED"):
        return s
    return CAT_ALIASES.get(s, CAT_ALIASES.get(s + ".", s))


#: How the workbook spells "riserva" in a free-text cell. Not a label: it is a
#: value read off the federation's file, matched here and written nowhere.
WORKBOOK_RESERVE = "RISERVA"


def parse_flag(value: Any) -> EventEntry | None:
    """`X` / `R` / `A` / `AR` / `1` -> EventEntry, None when not enrolled.

    A letter or a number is the pairing within the region - the madison coppia
    or squadra A/B of a team event. The same letter followed by `R` is that
    squadra's riserva: `A` rides, `AR` sits ready. `RA` and `R1` are the same
    thing written the other way round, which is how some regional forms spell
    it. All the notations are read; the app writes the letter.
    """
    s = _s(value).upper()
    if not s:
        return None
    if s in ("X", "SI", "S"):
        return EventEntry(starter=True)
    if len(s) == 2 and s.endswith(FLAG_RESERVE):  # "AR", "2R"
        n = pairing_number(s[0])
        if n is not None:
            return EventEntry(starter=False, pair=n)
    if len(s) == 2 and s.startswith(FLAG_RESERVE):  # "RA", "R1"
        n = pairing_number(s[1])
        if n is not None:
            return EventEntry(starter=False, pair=n)
    if s.startswith(FLAG_RESERVE):  # "R", "X RISERVA", "RISERVA"
        return EventEntry(starter=False, note=s if s != "R" else "")
    if WORKBOOK_RESERVE in s:
        return EventEntry(starter=False, note=s)
    n = pairing_number(s)
    if n is not None:
        return EventEntry(starter=True, pair=n)
    return EventEntry(starter=True, note=msg("xls_unknown_flag", value=repr(s)))


# ── import: master workbook ─────────────────────────────────────────────────

def _norm(text: Any) -> str:
    """Header key: case, dots and spaces do not count."""
    return "".join(_s(text).upper().split()).replace(".", "")


def _events_by_header(comp: Competition) -> dict[str, str]:
    """Workbook column header -> event code (case/space insensitive)."""
    out: dict[str, str] = {}
    for code, ev in comp.events.items():
        for name in [*ev.entry_columns, ev.name, ev.short, code]:
            if key := _norm(name):
                out.setdefault(key, code)
    return out


def _match_event(header: str, table: dict[str, str]) -> str | None:
    return table.get(_norm(header))


def sheet_names(path: str | Path) -> list[str]:
    import pandas as pd

    with pd.ExcelFile(path) as xls:
        return [str(n) for n in xls.sheet_names]


def is_flat_export(path: str | Path, comp: Competition) -> bool:
    """True when the file has no per-category sheets to read.

    The two shapes are told apart by what is in the workbook, not by its name:
    the master carries a sheet per categoria (`ES`, `ED`, `AL`, `DA`), the
    ksport export carries one flat list and nothing else.
    """
    names = {n.upper() for n in sheet_names(path)}
    return not names & {c.upper() for c in comp.cat_order()}


def import_entries(path: str | Path, comp: Competition) -> EntryList:
    """Read whichever entry file this is - the one call the app makes."""
    if is_flat_export(path, comp):
        return import_ksport_export(path, comp)
    return import_master(path, comp)


def import_master(path: str | Path, comp: Competition) -> EntryList:
    """Read the master ISCRITTI workbook into an EntryList."""
    import openpyxl

    path = Path(path)
    wb = openpyxl.load_workbook(path, data_only=True, read_only=False)
    el = EntryList(source_file=str(path), source_hash=file_hash(path),
                   imported_at=datetime.now().isoformat(timespec="seconds"))
    headers = _events_by_header(comp)
    sheet = comp.entry_sheet
    prov2reg = _read_province(wb)

    for cat in comp.cat_order():
        if cat not in wb.sheetnames:
            el.warnings.append(msg("xls_sheet_missing", sheet=cat))
            continue
        _read_category_sheet(wb[cat], cat, el, headers, sheet)

    if "KSPORT" in wb.sheetnames:
        _enrich_from_ksport(wb["KSPORT"], el, sheet, prov2reg)
    else:
        el.warnings.append(msg("xls_no_ksport"))

    wb.close()
    build_teams_and_pairs(el, comp)
    return el


def _read_province(wb) -> dict[str, str]:
    if "PROVINCE" not in wb.sheetnames:
        return {}
    ws = wb["PROVINCE"]
    out = {}
    for row in ws.iter_rows(values_only=True):
        if row and _s(row[0]) and len(row) > 2:
            out[_s(row[0]).upper()] = norm_region(row[2])
    return out


def _read_category_sheet(ws, cat: str, el: EntryList, headers: dict[str, str],
                         sheet: EntrySheet) -> None:
    """Parse one printable category sheet, columns matched by their heading."""
    hdr = {c: _s(ws.cell(sheet.header_row, c).value)
           for c in range(1, ws.max_column + 1)}

    # fixed columns: found by name, so a shifted or reordered sheet still reads
    at: dict[str, int] = {}
    for c, h in hdr.items():
        name = sheet.field_of(h)
        if name:
            at.setdefault(name, c)
    for missing in [f for f in sheet.fields if f not in at]:
        el.warnings.append(msg("xls_column_missing", cat=cat,
                               column=sheet.header_of(missing),
                               row=sheet.header_row))

    # Event columns run from the first column after the last fixed one until the
    # first blank header; anything past that is a hidden helper column. The
    # check-in columns are added by the giuria wherever there is room - before
    # the specialità or after them - so they count as neither boundary.
    event_cols: dict[int, str] = {}
    fixed = [c for f, c in at.items() if f not in CHECK_IN_FIELDS]
    for c in range(max(fixed, default=len(sheet.fields)) + 1,
                   ws.max_column + 1):
        h = hdr.get(c, "")
        if not h:
            break
        if sheet.field_of(h):        # a column we already know by name
            continue
        code = _match_event(h, headers)
        if code:
            event_cols[c] = code
        else:
            el.warnings.append(msg("xls_unknown_event_column", cat=cat,
                                   column=repr(h)))

    def cell(row: int, fname: str):
        c = at.get(fname)
        return ws.cell(row, c).value if c else None

    for r in range(sheet.first_data_row, ws.max_row + 1):
        bib = _int(cell(r, "bib"))
        uci = _uci(cell(r, "uci_id"))
        last_name = _s(cell(r, "last_name"))
        # Sheets end with reserved slots carrying only a bib and the category,
        # then a totals row. Neither is a rider.
        if not uci and not last_name:
            continue
        key = rider_key(uci, cat, bib)
        if key in el.riders:
            el.warnings.append(msg("xls_duplicate_rider", cat=cat, row=r,
                                   name=last_name, key=key))
            continue

        rider = Rider(
            key=key, bib=bib, last_name=last_name,
            first_name=_s(cell(r, "first_name")),
            cat=norm_cat(cell(r, "cat")) or cat, uci_id=uci,
            nation=_s(cell(r, "nation")) or "ITA",
            club=_s(cell(r, "club")), club_code=_s(cell(r, "club_code")),
            region=norm_region(cell(r, "region")), source=f"{ws.title}!{r}",
            checked_in=_yes(cell(r, "checked_in")),
            not_starting=_yes(cell(r, "not_starting")),
        )
        for c, code in event_cols.items():
            e = parse_flag(ws.cell(r, c).value)
            if e is not None:
                rider.events[code] = e
                if e.note.startswith(_unknown_flag()):
                    el.warnings.append(msg("xls_bad_flag", cat=cat,
                                           name=rider.full_name, event=code,
                                           note=e.note))
        if not rider.region:
            el.warnings.append(msg("xls_missing_region", cat=cat,
                                   name=rider.full_name))
        el.riders[key] = rider


def _enrich_from_ksport(ws, el: EntryList, sheet: EntrySheet,
                        prov2reg: dict[str, str]) -> None:
    """Add federal data (FCI code, birth date, certificate, province) by UCI ID."""
    # column index per *field*, via the `entries.ksport` mapping
    at = {}
    for i, value in enumerate(next(ws.iter_rows(values_only=True))):
        name = sheet.field_of(value, ksport=True)
        if name:
            at.setdefault(name, i)

    def col(row, name):
        i = at.get(name)
        return row[i] if i is not None and i < len(row) else None

    seen = set()
    for r, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        uci = _uci(col(row, "uci_id"))
        if not uci:
            continue
        seen.add(uci)
        rider = el.riders.get(uci)
        if rider is None:
            el.warnings.append(msg("ksport_not_in_category",
                                   name=_s(col(row, "full_name")), uci=uci))
            continue
        rider.fci_code = _s(col(row, "fci_code"))
        rider.birth_date = _date(col(row, "birth_date"))
        rider.sex = _s(col(row, "sex"))
        rider.province = _s(col(row, "province")).upper()
        rider.note = _s(col(row, "note"))
        rider.reserve_entry = _s(col(row, "reserve_entry")).upper() \
            in ("SI", "S", "YES")
        rider.certificate_date = _date(col(row, "certificate_date"))
        rider.ksport_source = f"{ws.title}!{r}"
        # The licence check lives in both places when the giuria added the
        # columns to both: a tick anywhere counts, so a rider verified by hand
        # in the foglio di categoria is not undone by a blank KSPORT cell (and
        # the other way round). Untick from the app clears both at once.
        rider.checked_in = rider.checked_in or _yes(col(row, "checked_in"))
        rider.not_starting = rider.not_starting or _yes(col(row, "not_starting"))
        if not rider.region:
            rider.region = (norm_region(col(row, "region"))
                            or _resolve_region_from_note(rider.note)
                            or prov2reg.get(rider.province, ""))

    missing = [r for r in el.riders.values() if r.uci_id and r.uci_id not in seen]
    if missing:
        el.warnings.append(msg(
            "ksport_missing", n=len(missing),
            who=", ".join(f"{r.cat} {r.bib} {r.full_name}" for r in missing[:8]),
            more=" ..." if len(missing) > 8 else ""))


def _resolve_region_from_note(note: str) -> str:
    """'Iscrizione CR. VENETO' -> 'VENETO' (blank when the region is missing)."""
    s = _s(note)
    marker = "Iscrizione CR."
    if marker in s:
        return norm_region(s.split(marker, 1)[1])
    return ""


# ── import: alternative sources ─────────────────────────────────────────────

def import_regional_form(path: str | Path, comp: Competition,
                         el: EntryList | None = None) -> EntryList:
    """Read a regional `Tabella-specialita-Giovanili.xlsx` submission.

    These arrive heterogeneous: long-form categories with trailing spaces,
    UCI IDs stored as numbers, `X RISERVA` free text. Normalised here so the
    merge into the master list is mechanical.
    """
    import openpyxl

    path = Path(path)
    el = el or EntryList(source_file=str(path),
                         imported_at=datetime.now().isoformat(timespec="seconds"))
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    headers = _events_by_header(comp)
    sheet = comp.entry_sheet

    hdr = {c: _s(ws.cell(1, c).value) for c in range(1, ws.max_column + 1)}
    at = {name: c for c, h in hdr.items()
          if h and (name := sheet.field_of(h))}
    event_cols = {c: code for c, h in hdr.items()
                  if h and (code := _match_event(h, headers))}

    def cell(r, name):
        c = at.get(name)
        return ws.cell(r, c).value if c else None

    for r in range(2, ws.max_row + 1):
        uci = _uci(cell(r, "uci_id"))
        last_name = _s(cell(r, "last_name"))
        if not uci and not last_name:
            continue
        cat = norm_cat(cell(r, "cat"))
        key = rider_key(uci, cat, None)
        rider = el.riders.get(key) or Rider(key=key, cat=cat, uci_id=uci,
                                            source=f"{path.name}!{r}")
        rider.last_name = rider.last_name or last_name
        rider.first_name = rider.first_name or _s(cell(r, "first_name"))
        rider.nation = rider.nation or _s(cell(r, "nation")) or "ITA"
        rider.club = rider.club or _s(cell(r, "club"))
        rider.club_code = rider.club_code or _s(cell(r, "club_code"))
        rider.region = rider.region or norm_region(cell(r, "region"))
        if not rider.cat:
            el.warnings.append(msg("xls_unknown_cat", file=path.name, row=r,
                                   value=repr(_s(cell(r, "cat")))))
        for c, code in event_cols.items():
            e = parse_flag(ws.cell(r, c).value)
            if e is not None:
                rider.events[code] = e
        el.riders[key] = rider
    wb.close()
    return el


def import_ksport_export(path: str | Path, comp: Competition) -> EntryList:
    """Read a flat ksport `Iscritti_NNNNNN_KSPORT.xlsx` export.

    One row per rider and no event columns: this file says who is entered in
    the *competition*, not in which specialità. Those are the programme's
    (`comp.events_for(cat)`) and the jury's, ticked at the verifica and kept in
    the overlay - which is why re-importing this file never disturbs them.

    Nothing is inferred that the file does not say: the categoria is read from
    its own column, not from the sheet a rider sits on, and the squadra from
    the column that names it (`Regione` / `NomeSocieta`), falling back to the
    "Iscrizione CR. ..." note only when there is no column at all.
    """
    import pandas as pd

    path = Path(path)
    df = pd.read_excel(path, dtype=str)
    el = EntryList(source_file=str(path), source_hash=file_hash(path),
                   imported_at=datetime.now().isoformat(timespec="seconds"))
    sheet = comp.entry_sheet
    # the export's own headings, mapped through `entries.ksport` (plus the
    # sheet columns: this file spells Cognome / Nome the same way)
    field_of = {}
    for column in df.columns:
        name = (sheet.field_of(column, ksport=True) or sheet.field_of(column))
        if name:
            field_of[name] = column

    def val(row, name):
        column = field_of.get(name)
        return row.get(column) if column else None

    for i, row in df.iterrows():
        uci = _uci(val(row, "uci_id"))
        cat = norm_cat(val(row, "cat"))
        bib = _int(val(row, "bib"))
        last_name, first_name = _names(val(row, "last_name"),
                                       val(row, "first_name"),
                                       val(row, "full_name"))
        if not uci and not last_name:
            continue
        key = rider_key(uci, cat, bib)
        if key in el.riders:
            el.warnings.append(msg("xls_duplicate_rider", cat=cat, row=i + 2,
                                   name=last_name, key=key))
            continue
        note = _s(val(row, "note"))
        el.riders[key] = Rider(
            key=key, bib=bib, last_name=last_name, first_name=first_name,
            cat=cat, uci_id=uci, fci_code=_s(val(row, "fci_code")),
            nation=_s(val(row, "nation")) or "ITA",
            birth_date=_date(val(row, "birth_date")), club=_s(val(row, "club")),
            club_code=_s(val(row, "club_code")), sex=_s(val(row, "sex")),
            province=_s(val(row, "province")).upper(), note=note,
            # the note spells the regione short ("Iscrizione CR. EMILIA ROM"):
            # it answers only where the export has no column of its own
            region=(norm_region(val(row, "region"))
                    or _resolve_region_from_note(note)),
            reserve_entry=_s(val(row, "reserve_entry")).upper()
            in ("SI", "S", "YES"),
            certificate_date=_date(val(row, "certificate_date")),
            checked_in=_yes(val(row, "checked_in")),
            not_starting=_yes(val(row, "not_starting")),
            source=f"{path.name}!{i + 2}",
        )
        # a categoria the programme does not run: the rider would be on no
        # sheet at all, so it is said out loud rather than left to be noticed
        if cat not in comp.categories:
            el.warnings.append(msg("xls_unknown_cat", file=path.name, row=i + 2,
                                   value=repr(_s(val(row, "cat")))))
    _report_missing_columns(el, field_of, comp)
    build_teams_and_pairs(el, comp)
    return el


#: What the flat export has to name for the app to work with it at all.
FLAT_REQUIRED = ("uci_id", "last_name", "cat", "bib")


def _report_missing_columns(el: EntryList, found: dict, comp: Competition) -> None:
    """Say which of the columns we need this export does not have."""
    sheet = comp.entry_sheet
    for name in [*FLAT_REQUIRED, comp.team_group]:
        if name not in found:
            el.warnings.append(msg(
                "flat_column_missing", field=label(name),
                header=sheet.header_of(name, ksport=True)))


def _names(last: Any, first: Any, full: Any) -> tuple[str, str]:
    """Cognome and Nome, from their own columns or split off `NomeTesserato`."""
    last, first = _s(last), _s(first)
    if last or first:
        return last, first
    parts = _s(full).split()
    return (" ".join(parts[:-1]), parts[-1]) if len(parts) > 1 else (
        " ".join(parts), "")


# ── derived entities: teams and madison pairs ───────────────────────────────

def build_teams_and_pairs(el: EntryList, comp: Competition) -> None:
    """(Re)build Team and Pair objects from the riders' event flags."""
    el.teams.clear()
    el.pairs.clear()
    el.errors.clear()  # they are re-stated below, from the flags as they are now
    for code, sp in comp.events.items():
        if sp.fmt == "timed_team":
            _build_teams(el, comp, code, sp.team_size or 4)
        elif sp.fmt == "madison":
            _build_pairs(el, comp, code)


def _by_region(riders: Iterable[Rider],
               merge: dict[str, str] | None = None) -> dict[str, list[Rider]]:
    """Riders by regione, in bib order.

    `merge` puts two rappresentative in the same bucket under one name, where
    they were authorised to field a squadra together (`Competition.team_merge`).
    Passed only where the squadre are composed: everything else - the quote
    above all - counts each regione as itself.
    """
    out: dict[str, list[Rider]] = {}
    for r in sorted(riders, key=lambda r: (r.bib or 9999, r.last_name)):
        region = r.region or "?"
        out.setdefault((merge or {}).get(region, region), []).append(r)
    return out


def _build_teams(el: EntryList, comp: Competition, event: str, size: int) -> None:
    """The teams of each region, as the jury wrote them at the check-in.

    A letter (`A`, `B`, `C`, ...) pins a rider to that squadra, and the same
    letter with an `R` is its riserva: four `A` and one `AR` make squadra A.
    Plain `X` says only "in this event": it is the whole squadra while a region
    fields one, and a rider still to be assigned once the letters are out.

    Teams are never composed here - splitting a region in bib order would
    invent a quartetto the jury never decided. A lettered squadra that does not
    field exactly `size` riders is an error: the jury wrote it, so it has to
    be right. An `X` region is still provisional and only warns.

    Two rappresentative authorised to ride together (`entries.team_merge`) are
    one region here, under the name of the joint squadra.
    """
    for cat in comp.cat_order():
        pool = [r for r in el.riders.values()
                if r.cat == cat and event in r.events and not r.not_starting]
        for region, riders in _by_region(pool, comp.team_merge(event)).items():
            lettered: dict[int, list[Rider]] = {}
            spare: dict[int, list[Rider]] = {}   # AR, BR, ...: per squadra
            loose: list[Rider] = []              # X, still to be assigned
            free: list[Rider] = []               # R alone: riserva of the region
            for r in riders:
                e = r.events[event]
                if e.pair:
                    (lettered if e.starter else spare).setdefault(
                        e.pair, []).append(r)
                else:
                    (loose if e.starter else free).append(r)

            where = f"[{cat} {comp.event(event).short}] {region}"
            if lettered and loose:
                el.errors.append(msg("team_loose_x", where=where,
                                     n=len(loose), bibs=_bibs(loose),
                                     letter=pairing_letter(1)))

            teams = sorted(lettered.items()) or \
                ([(None, loose)] if loose or free else [])
            multi = len(teams) > 1
            for n, chunk in teams:
                letter = pairing_letter(n) if multi else ""
                res = spare.get(n, []) + (free if n is None else [])
                key = f"{cat}:{event}:{region}:{letter}"
                el.teams[key] = Team(key=key, cat=cat, event=event,
                                     region=region, letter=letter,
                                     riders=[r.key for r in chunk],
                                     reserves=[r.key for r in res])
                who = (msg("team_letter", where=where, letter=letter)
                       if letter else where)
                if n is not None:
                    if len(chunk) != size:
                        el.errors.append(msg("team_wrong_size", where=who,
                                             n=len(chunk), size=size,
                                             bibs=_bibs(chunk)))
                # the provisional squadra of a region that has not been split
                # yet: one over the size is the riserva written as an X
                elif chunk and not size <= len(chunk) <= size + 1:
                    el.errors.append(msg(
                        "team_region_wrong_size", where=who, n=len(chunk),
                        size=size, bibs=_bibs(chunk),
                        hint=(msg("team_compose_hint")
                              if len(chunk) > size else ".")))


def _bibs(riders: list[Rider]) -> str:
    return ", ".join(str(r.bib) for r in riders if r.bib is not None)


def _build_pairs(el: EntryList, comp: Competition, event: str) -> None:
    """Group riders by (region, pair number). Riders flagged `X` are unpaired."""
    for cat in comp.cat_order():
        pool = [r for r in el.riders.values()
                if r.cat == cat and event in r.events and not r.not_starting]
        for region, riders in _by_region(pool, comp.team_merge(event)).items():
            numbered: dict[int, list[Rider]] = {}
            unnumbered: list[Rider] = []
            reserves: list[Rider] = []
            for r in riders:
                e = r.events[event]
                if not e.starter:
                    reserves.append(r)
                elif e.pair is not None:
                    numbered.setdefault(e.pair, []).append(r)
                else:
                    unnumbered.append(r)
            # riders with a bare X are paired in bib order, after the numbered ones
            n = max(numbered) if numbered else 0
            for i in range(0, len(unnumbered), 2):
                n += 1
                numbered[n] = unnumbered[i:i + 2]
            # a region with one coppia is just the region; with two they are
            # A and B, the same way its quartetti are
            multi = len(numbered) > 1
            for num, members in sorted(numbered.items()):
                letter = pairing_letter(num) if multi else ""
                key = f"{cat}:{region}:{num}"
                el.pairs[key] = Pair(key=key, cat=cat, region=region, number=num,
                                     letter=letter,
                                     riders=[r.key for r in members],
                                     reserves=[r.key for r in reserves])
                if len(members) != 2:
                    el.errors.append(msg(
                        "pair_wrong_size", cat=cat,
                        event=comp.event(event).short,
                        who=f"{region} {letter}".rstrip(), n=len(members)))


# ── overlay of jury edits ───────────────────────────────────────────────────

@dataclass
class Patch:
    """One explicit, reversible edit made in the app on top of the import."""

    target: str                  # rider key
    op: str                      # one of OPS
    field: str = ""
    value: Any = None
    reason: str = ""
    ts: str = ""
    actor: str = ""

    def __post_init__(self):
        self.ts = self.ts or datetime.now().isoformat(timespec="seconds")


def _unknown_flag() -> str:
    """Prefix of the note `parse_flag` leaves on a workbook value it cannot read.

    Matched again on import to report it, so the two never drift apart - and
    read from the catalogue each time rather than frozen at import, or a
    workbook read after a change of language would match nothing.
    """
    return msg("xls_unknown_flag", value="").rstrip("'\" ")

OPS = ("set_field", "set_checked_in", "set_not_starting", "set_event",
       "clear_event", "set_pair")

# Ticking a rider in at the licence desk needs no written reason; every other
# edit does. See ui/pages/verifica.py.
CHECK_IN_OPS = ("set_checked_in", "set_not_starting")

#: The same two, as rider fields. The federation's file has no column for them:
#: they are written back only where the giuria added one (`entries.check_in`).
CHECK_IN_FIELDS = ("checked_in", "not_starting")


def check_in_columns(comp: Competition) -> tuple[str, ...]:
    """Which of the two the entry file is declared to have a column for."""
    declared = set(comp.entry_sheet.check_in.values())
    return tuple(f for f in CHECK_IN_FIELDS if f in declared)

# Names used before the code moved to English, still readable from an overlay
# written by an earlier version.
LEGACY_OPS = {"set_np": "set_not_starting", "set_verificato": "set_checked_in",
              "set_spec": "set_event", "clear_spec": "clear_event"}


def apply_overlay(el: EntryList, patches: list[Patch], comp: Competition
                  ) -> list[str]:
    """Apply patches in order; returns the ones that no longer apply."""
    stale: list[str] = []
    for p in patches:
        rider = el.riders.get(p.target)
        if rider is None:
            stale.append(msg("patch_rider_gone", op=p.op, target=p.target,
                             reason=p.reason))
            continue
        op = LEGACY_OPS.get(p.op, p.op)
        field = LEGACY_KEYS.get(p.field, p.field)
        if op == "set_field":
            if not hasattr(rider, field):
                stale.append(msg("patch_unknown_field", field=p.field,
                                 target=p.target))
                continue
            setattr(rider, field, p.value)
        elif op == "set_not_starting":
            rider.not_starting = bool(p.value)
        elif op == "set_checked_in":
            rider.checked_in = bool(p.value)
        elif op == "set_event":
            e = parse_flag(p.value) if not isinstance(
                p.value, dict) else EventEntry(**p.value)
            if e is None:
                rider.events.pop(p.field, None)
            else:
                rider.events[p.field] = e
        elif op == "clear_event":
            if rider.events.pop(p.field, None) is None:
                stale.append(msg("patch_not_entered", target=p.target,
                                 event=p.field))
        elif op == "set_pair":
            e = rider.events.get(p.field)
            if e is None:
                stale.append(msg("patch_no_pair_entry", target=p.target,
                                 event=p.field))
                continue
            e.pair = _int(p.value)
        else:
            stale.append(msg("patch_unknown_op", op=p.op, target=p.target))
    build_teams_and_pairs(el, comp)
    return stale


def patches_to_json(patches: list[Patch]) -> list[dict]:
    return [asdict(p) for p in patches]


def patches_from_json(data: Any) -> list[Patch]:
    return [Patch(**d) for d in (data or [])]


# ── persistence: import snapshot + overlay ──────────────────────────────────

IMPORT_FILE = "entries_import.json"
OVERLAY_FILE = "entries_overlay.json"


#: Where the entry file is, when this machine keeps its own copy of it (a Drive
#: folder that is not the one the programme was written with, a USB stick).
SOURCE_SETTING = "entries_source"


def source_path(store, comp: Competition) -> str:
    """The entry file to read: the one chosen in Impostazioni, else the programme's."""
    return str(store.settings.get(SOURCE_SETTING) or comp.entries_source or "")


def set_source_path(store, path: str) -> None:
    store.set_setting(SOURCE_SETTING, str(path).strip())


def save_import(store, el: EntryList) -> None:
    store.write_json(IMPORT_FILE, el.to_dict(), action="import_entries")


def load_import(store) -> EntryList | None:
    d = store.read_json(IMPORT_FILE)
    return EntryList.from_dict(d) if d else None


def load_overlay(store) -> list[Patch]:
    return patches_from_json(store.read_json(OVERLAY_FILE, []))


def save_overlay(store, patches: list[Patch], action: str = "edit_entries") -> None:
    store.write_json(OVERLAY_FILE, patches_to_json(patches), action=action)


#: Whether the jury's edits are applied on top of the import at all. On by
#: default, and it is what the Verifica page is for. Off, the entry list *is*
#: the workbook: the file is the one place the entries are changed, and it is
#: re-imported. The overlay is not deleted - it is set aside, and comes back
#: whole the moment the setting goes back on.
OVERLAY_SETTING = "use_overlay"


def overlay_on(store) -> bool:
    """Whether the overlay of jury edits is in force (Impostazioni)."""
    return store.settings.get(OVERLAY_SETTING, True) is not False


def set_overlay_on(store, on: bool) -> None:
    # written even when it is False: `set_setting` drops None and "", not False
    store.set_setting(OVERLAY_SETTING, bool(on))


def effective_entries(store, comp: Competition, *, overlay: bool | None = None
                      ) -> tuple[EntryList | None, list[str]]:
    """Import snapshot + overlay applied. Returns (list, stale-patch messages).

    `overlay` overrides the setting, for a caller that has to see one or the
    other whatever the competition is set to.
    """
    el = load_import(store)
    if el is None:
        return None, []
    if not (overlay_on(store) if overlay is None else overlay):
        # the squadre and the coppie are read off the entries, and applying the
        # overlay is what normally builds them: with no overlay to apply they
        # still have to be built
        build_teams_and_pairs(el, comp)
        return el, []
    stale = apply_overlay(el, load_overlay(store), comp)
    return el, stale


def source_changed(store, path: str | Path) -> bool:
    """True when the master workbook has been modified since the last import."""
    el = load_import(store)
    if el is None or not Path(path).exists():
        return el is None
    return el.source_hash != file_hash(path)


# ── validation ──────────────────────────────────────────────────────────────

def _minus_one_year(iso_date: str) -> str:
    """'2026-08-04' -> '2025-08-04' (blank if not an ISO date)."""
    try:
        y, m, d = iso_date.split("-")
        return f"{int(y) - 1:04d}-{m}-{d}"
    except (ValueError, AttributeError):
        return ""


@dataclass
class CheckInProgress:
    """How far the licence check has got, for one category or the whole competition."""

    entries: int = 0  # riders on the list, NP excluded
    verificati: int = 0  # ticked as present at the licence check
    missing: int = 0  # neither verified nor declared NP
    not_starting: int = 0  # declared non partenti

    @property
    def done(self) -> bool:
        return self.missing == 0 and self.entries > 0


def check_in_progress(el: EntryList, cat: str = "") -> CheckInProgress:
    riders = el.by_cat(cat) if cat else list(el.riders.values())
    p = CheckInProgress()
    for r in riders:
        if r.not_starting:
            p.not_starting += 1
            continue
        p.entries += 1
        if r.checked_in:
            p.verificati += 1
        else:
            p.missing += 1
    return p


def guessed_pairings(el: EntryList, comp: Competition
                     ) -> list[tuple[str, str, str, list[Rider]]]:
    """Where the madison coppie were guessed instead of declared.

    A rider entered with a bare `X` says "rides the madison", not *with whom*:
    `_build_pairs` then pairs them two by two in bib order. With two of them
    there is only one way to do it, so nothing is really guessed - from three
    up the app has picked an accoppiamento the jury never wrote down, and it
    has to confirm it before the numbers go out.

    Returns (cat, event, region, riders) per region concerned.
    """
    out = []
    for event in comp.event_order():
        if comp.event(event).fmt != "madison":
            continue
        for cat in comp.cat_order():
            pool = [r for r in el.riders.values()
                    if r.cat == cat and event in r.events and not r.not_starting]
            for region, riders in _by_region(pool).items():
                loose = [r for r in riders
                         if r.events[event].starter
                         and r.events[event].pair is None]
                if len(loose) > 2:
                    out.append((cat, event, region,
                                sorted(loose, key=lambda r: (r.bib is None,
                                                             r.bib or 0))))
    return out


def validate_entries(el: EntryList, comp: Competition) -> list[Issue]:
    """Quota and consistency checks. Warnings only - nothing here blocks work."""
    issues: list[Issue] = [Issue(ERROR, "teams", e) for e in el.errors]
    issues += [Issue(WARN, "import", w) for w in el.warnings]
    q = comp.quotas

    for cat, event, region, riders in guessed_pairings(el, comp):
        issues.append(Issue(WARN, "pairs", msg(
            "pairs_guessed", cat=cat, event=comp.event(event).short,
            region=region, n=len(riders), bibs=_bibs(riders))))

    for r in el.riders.values():
        if r.not_starting:
            continue
        if not r.uci_id:
            issues.append(Issue(ERROR, "uci", msg(
                "rider_no_uci", cat=r.cat, bib=r.bib, name=r.full_name), r.key))
        # what a rider rides *for*: the regione at a championship, the società
        # at an open meeting - whichever this competition groups by
        if not getattr(r, comp.team_group, ""):
            issues.append(Issue(ERROR, "region",
                                msg("rider_no_team", name=r.full_name,
                                    what=label(comp.team_group)), r.key))
        if r.bib is None:
            issues.append(Issue(ERROR, "bib",
                                msg("rider_no_bib", name=r.full_name), r.key))
        # an entry the programme does not run: with the flat export the
        # specialità are ticked by hand, and a tick in the wrong column would
        # otherwise be found only when the sheet comes out empty
        run = comp.events_for(r.cat)
        for s in r.events:
            if s != EVENT_ENTRY_LIST and s not in run:
                issues.append(Issue(WARN, "event_not_run", msg(
                    "rider_event_not_run", cat=r.cat, bib=r.bib,
                    name=r.full_name, event=comp.event(s).short), r.key))
        lim = q.max_events_per_rider.get(r.cat)
        if lim and q.max_events_level != "off":
            n = r.n_events(include_reserves=q.max_events_count_reserves)
            if n > lim:
                events = ", ".join(comp.event(s).short
                                   for s in comp.event_order()
                                   if s in r.events and s != EVENT_ENTRY_LIST)
                issues.append(Issue(q.max_events_level, "quota_rider", msg(
                    "rider_over_events", cat=r.cat, bib=r.bib,
                    name=r.full_name, n=n, max=lim, events=events), r.key))
        # The ksport column is the certificate's *issue* date, not its expiry:
        # in the 2026 export not one of the 227 dates falls after the competition.
        # A medical certificate lasts one year, so flag anything older than that.
        if r.certificate_date and comp.dates:
            limit = _minus_one_year(comp.dates[0])
            if limit and r.certificate_date < limit:
                issues.append(Issue(WARN, "certificate", msg(
                    "rider_old_certificate", cat=r.cat, bib=r.bib,
                    name=r.full_name, date=r.certificate_date), r.key))

    # duplicate bibs within a category
    for cat in comp.cat_order():
        seen: dict[int, Rider] = {}
        for r in el.by_cat(cat):
            if r.bib is None:
                continue
            if r.bib in seen:
                issues.append(Issue(ERROR, "bib_dup", msg(
                    "duplicate_bib", cat=cat, bib=r.bib,
                    a=seen[r.bib].full_name, b=r.full_name)))
            seen[r.bib] = r

    # per-region and per-club quotas
    for cat in comp.cat_order():
        for event in comp.events_for(cat):
            entered = el.entered(cat, event)
            lim = q.max_per_region.get(event)
            if lim:
                for region, riders in _by_region(entered).items():
                    if len(riders) > lim:
                        issues.append(Issue(WARN, "quota_region", msg(
                            "quota_region", cat=cat,
                            event=comp.event(event).short, region=region,
                            n=len(riders), max=lim)))
            lim = q.max_same_club_per_region.get(event)
            if lim:
                per: dict[tuple[str, str], list[Rider]] = {}
                for r in entered:
                    per.setdefault((r.region, r.club), []).append(r)
                for (region, club), riders in per.items():
                    if len(riders) > lim and club:
                        issues.append(Issue(WARN, "quota_club_region", msg(
                            "quota_club_region", cat=cat,
                            event=comp.event(event).short, region=region or "?",
                            club=club, n=len(riders), max=lim,
                            bibs=_bibs(riders))))
            lim = q.max_same_club.get(event)
            if lim:
                clubs: dict[str, int] = {}
                for r in entered:
                    clubs[r.club] = clubs.get(r.club, 0) + 1
                for club, n in clubs.items():
                    if n > lim and club:
                        issues.append(Issue(WARN, "quota_club", msg(
                            "quota_club", cat=cat,
                            event=comp.event(event).short, club=club,
                            n=n, max=lim)))

    # teams and pairs per region
    for event, lim in q.max_teams_per_region.items():
        groups: dict[tuple[str, str], int] = {}
        if comp.event(event).fmt == "madison":
            for p in el.pairs.values():
                groups[(p.cat, p.region)] = groups.get((p.cat, p.region), 0) + 1
        else:
            for t in el.teams.values():
                if t.event == event and t.riders:
                    groups[(t.cat, t.region)] = groups.get(
                        (t.cat, t.region), 0) + 1
        for (cat, region), n in groups.items():
            if n > lim:
                issues.append(Issue(WARN, "quota_teams", msg(
                    "quota_teams", cat=cat, event=comp.event(event).short,
                    region=region, n=n, max=lim)))
    return issues


# ── writing the entries back into the workbook ──────────────────────────────
#
# With the overlay off (`OVERLAY_SETTING`) the workbook is the master and the
# app edits it: what the jury types in Verifica is written into the file
# itself, which is then re-imported. Only what the file has a column for can be
# written, and the previous version is always copied aside first, because this
# is the one file the app does not own.

#: The edits a workbook can hold. The licence check is written too, but only
#: where a column for it exists (`entries.check_in`): the federation's layout
#: has none, so it is the giuria that adds it - see `_write_check_in`.
WRITABLE_OPS = ("set_field", "set_event", "clear_event", "set_pair")

#: Where the copy of the file made before writing to it goes, under the
#: competition folder - never next to the original, which may be a Drive folder
#: the federation also writes to.
SOURCE_BACKUP = "entries_source"


def backup_source(store, path: str | Path) -> Path:
    """Copy the entry workbook aside, under the competition's snapshots."""
    import shutil

    path = Path(path)
    folder = Path(store.root) / ".snapshots" / SOURCE_BACKUP
    folder.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    dst = folder / f"{stamp}_{path.name}"
    n = 0
    while dst.exists():
        n += 1
        dst = folder / f"{stamp}-{n}_{path.name}"
    shutil.copy2(path, dst)
    store.journal(action="backup_entries_source", target=str(path))
    return dst


def write_back(path: str | Path, comp: Competition, el: EntryList,
               patches: list[Patch], *, store=None) -> tuple[int, list[str]]:
    """Write the jury's edits into the entry workbook itself.

    `el` is the list as imported: the patches are applied to it and every cell
    they touch is written with the *resulting* value, so the file ends up
    saying exactly what the app shows. Returns how many cells were written and
    the edits the file has no place for, each with its reason.

    The rows are found through `Rider.source` - the sheet and row each rider
    was read from - and checked against the UCI ID before anything is written:
    a file edited by hand since the import must not have the wrong line
    overwritten.
    """
    import openpyxl

    path = Path(path)
    if path.suffix.lower() != ".xlsx":
        # openpyxl writes .xlsx only, and a silent no-op here would look like
        # a save that worked
        return 0, [msg("write_back_not_xlsx", file=path.name)]

    refused = list(apply_overlay(el, patches, comp))
    flat = is_flat_export(path, comp)
    if store is not None:
        backup_source(store, path)

    # not `data_only`: read that way, every formula in the workbook would be
    # saved back as the value it last held
    wb = openpyxl.load_workbook(path)
    headers = _events_by_header(comp)
    maps: dict[str, tuple[dict[str, int], dict[str, int]]] = {}
    written = 0

    for p in patches:
        rider = el.riders.get(p.target)
        if rider is None:
            continue                      # apply_overlay already said so
        op = LEGACY_OPS.get(p.op, p.op)
        if op in CHECK_IN_OPS:
            n, why = _write_check_in(wb, rider, _op_field(op),
                                     comp.entry_sheet, flat, headers, maps)
            written += n
            refused += why
            continue
        if op not in WRITABLE_OPS:
            refused.append(msg("write_back_no_column", name=rider.full_name,
                               what=label(LEGACY_KEYS.get(p.field, p.field)
                                          or _op_field(op))))
            continue
        found = _locate_row(wb, rider, comp.entry_sheet, flat, headers, maps)
        if found is None:
            refused.append(msg("write_back_row_gone", name=rider.full_name,
                               source=rider.source or "?"))
            continue
        ws, row, at, event_cols = found

        if op == "set_field":
            field = LEGACY_KEYS.get(p.field, p.field)
            col = at.get(field)
            if col is None:
                refused.append(msg("write_back_no_column",
                                   name=rider.full_name, what=label(field)))
                continue
            value = getattr(rider, field, None)
        else:
            code = p.field
            col = event_cols.get(code)
            if col is None:
                refused.append(msg("write_back_no_event_column",
                                   name=rider.full_name,
                                   event=comp.event(code).short))
                continue
            e = rider.events.get(code)
            value = e.flag if e else None

        ws.cell(row, col).value = value if value not in ("", None) else None
        written += 1

    if written:
        wb.save(path)
    wb.close()
    return written, refused


def _op_field(op: str) -> str:
    """The field name behind an op that does not carry one."""
    return {"set_checked_in": "checked_in",
            "set_not_starting": "not_starting"}.get(op, op)


def _write_check_in(wb, rider: Rider, field: str, sheet: EntrySheet, flat: bool,
                    headers: dict[str, str], maps: dict) -> tuple[int, list[str]]:
    """Tick verificato / NP in every sheet that has a column for it.

    The master workbook carries the same rider twice - on the foglio di
    categoria the giuria prints and on the KSPORT sheet the federation sent -
    so both are written and a re-import reads the same answer from either one.
    Written as SI / blank, the way the workbook writes its own yes-no columns.
    """
    value = YES if getattr(rider, field, False) else None
    rows = [r for r in (_locate_row(wb, rider, sheet, flat, headers, maps),
                        _locate_ksport_row(wb, rider, sheet, maps))
            if r is not None]
    if not rows:
        return 0, [msg("write_back_row_gone", name=rider.full_name,
                       source=rider.source or "?")]
    written = 0
    for ws, row, at, _ in rows:
        col = at.get(field)
        if col is None:
            continue
        ws.cell(row, col).value = value
        written += 1
    if not written:
        return 0, [msg("write_back_no_column", name=rider.full_name,
                       what=label(field))]
    return written, []


def _locate_ksport_row(wb, rider: Rider, sheet: EntrySheet, maps: dict):
    """The rider's row on the KSPORT sheet, checked against the UCI ID."""
    name, _, r = (rider.ksport_source or "").rpartition("!")
    row = _int(r)
    if not row or name not in wb.sheetnames:
        return None
    ws = wb[name]
    if row > ws.max_row:
        return None
    if ws.title not in maps:
        # the KSPORT sheet reads by the `ksport` mapping and heads its first
        # row, exactly like the flat export
        maps[ws.title] = _sheet_columns(ws, sheet, True, {})
    at, event_cols = maps[ws.title]
    col = at.get("uci_id")
    if not col or not rider.uci_id:
        return None
    if _uci(ws.cell(row, col).value) != rider.uci_id:
        return None
    return ws, row, at, event_cols


def _locate_row(wb, rider: Rider, sheet: EntrySheet, flat: bool,
                headers: dict[str, str], maps: dict):
    """(worksheet, row, fields->col, events->col) for one rider, or None."""
    name, _, r = (rider.source or "").rpartition("!")
    row = _int(r)
    ws = (wb[name] if name in wb.sheetnames
          else wb.worksheets[0] if flat and wb.worksheets else None)
    if ws is None or not row or row > ws.max_row:
        return None
    if ws.title not in maps:
        maps[ws.title] = _sheet_columns(ws, sheet, flat, headers)
    at, event_cols = maps[ws.title]

    # the row has to still be this rider: the file may have been edited, or
    # rows inserted, since the import it was read from
    uci_col, name_col = at.get("uci_id"), at.get("last_name")
    if rider.uci_id and uci_col:
        if _uci(ws.cell(row, uci_col).value) != rider.uci_id:
            return None
    elif name_col:
        if _s(ws.cell(row, name_col).value).upper() != rider.last_name.upper():
            return None
    else:
        return None
    return ws, row, at, event_cols


def _sheet_columns(ws, sheet: EntrySheet, flat: bool, headers: dict[str, str]
                   ) -> tuple[dict[str, int], dict[str, int]]:
    """Which column holds which field, and which one holds which event.

    The same rule the import reads by: matched on the heading, never on the
    position, so a workbook with a column inserted still writes where it reads.
    """
    header_row = 1 if flat else sheet.header_row
    at: dict[str, int] = {}
    event_cols: dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        h = _s(ws.cell(header_row, c).value)
        if not h:
            continue
        field = (sheet.field_of(h, ksport=True) or sheet.field_of(h) if flat
                 else sheet.field_of(h))
        if field:
            at.setdefault(field, c)
        elif not flat:
            code = _match_event(h, headers)
            if code:
                event_cols.setdefault(code, c)
    return at, event_cols


# ── export ──────────────────────────────────────────────────────────────────

def export_xlsx(el: EntryList, comp: Competition, path: str | Path) -> Path:
    """Write the effective list in the FCI submission layout, one sheet per category."""
    import openpyxl

    path = Path(path)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for cat in comp.cat_order():
        riders = sorted(el.by_cat(cat), key=lambda r: (r.bib or 9999, r.last_name))
        if not riders:
            continue
        events = [s for s in (comp.events_for(cat) or comp.event_order())
                  if s != EVENT_ENTRY_LIST]
        sheet = comp.entry_sheet
        ws = wb.create_sheet(cat)
        # written back in the workbook's own wording, from `entries.columns`
        ws.append([sheet.header_of(f) for f in sheet.fields]
                  + [comp.event(s).short for s in events])
        for r in riders:
            values = {**{f: getattr(r, f, "") for f in sheet.fields},
                      "cat": "NP" if r.not_starting else r.cat}
            row = [values[f] for f in sheet.fields]
            for s in events:
                e = r.events.get(s)
                row.append(e.flag if e else "")
            ws.append(row)
    wb.save(path)
    return path
