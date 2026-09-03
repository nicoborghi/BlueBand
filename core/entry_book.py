"""The elenco iscritti of *this* competition, built from what the federation sends.

What arrives is a flat export: one row per rider, a categoria beside each, and
nothing about which event anybody rides - because the federation does not
know, it is the programme that says which events exist and the giuria that
says who is in them. What the competition is then run from is a workbook with a
sheet per categoria and a column per event of that categoria, which is
where the X marks go, plus the federal export kept whole on a sheet of its own.

Until now that workbook was made by hand, once, before every championship. This
module makes it:

    build(entries, comp, path)   the workbook, from an imported list
    sync(path, comp)             the same workbook, after the programme moved
    merge(old, new)              a corrected file, over the work already done
    numbered(entries, comp, how) dorsali, when the export has none

It is always the same file, `entry_list.xlsx`, in the folder of the
competition: one name, because the folder already says which meeting it is.

**The programme comes first.** A sheet per categoria with a column per
event cannot be written before somebody has said which categorie ride and
what each of them rides, which is why the page that calls this refuses to run
until the programme says so.

**The federal sheet is kept whole.** `_KSPORT` is the export as it arrived, with
two columns added - *Verificato* and *NP*, the two the giuria fills in and the
federation has no place for. Keeping it means a re-import can be checked
against what was actually received, and it is the sheet `entries.import_master`
enriches the categorie from. The underscore is the only thing that says so out
loud: it is an archive, and the fogli the giuria works on are the categorie.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path

from . import entries as E
from .config import EVENT_ENTRY_LIST, Competition

#: The sheet the federal export is kept on, whole. The underscore says what it
#: is: an archive of what arrived, not a foglio anybody works on - the categorie
#: are the sheets the giuria fills in. `entries` reads it under either name.
KSPORT = E.ARCHIVE_SHEET

#: What the workbook of a competition is called. One name for every
#: competition - the folder already says which one it is.
FILENAME = "entry_list.xlsx"

#: What it used to be called, as the federation names its own files. A folder
#: that has one and no `entry_list.xlsx` is still opened from it (`book_path`).
PREFIX = "Iscritti_"

#: How the dorsali are dealt out when the export has none.
AS_IMPORTED = "as_imported"      # 1..N in the order the file lists them
BY_CAT_RUNNING = "by_cat"        # 1..N, N+1..M: one run, categoria by categoria
BY_CAT_RESTART = "by_cat_restart"  # 1..N, 1..M: each categoria from 1
NUMBERINGS = (AS_IMPORTED, BY_CAT_RUNNING, BY_CAT_RESTART)


def book_path(root: str | Path) -> Path:
    """The workbook of the competition in that folder.

    `entry_list.xlsx` if it is there or if the folder is empty of entry files;
    the old `Iscritti_*.xlsx` when that is what the folder has, so a
    competition set up before the rename opens on the file it already owns
    instead of on an import page saying there is nothing.
    """
    root = Path(root)
    current = root / FILENAME
    if current.exists():
        return current
    legacy = sorted(root.glob(f"{PREFIX}*.xlsx"))
    return legacy[0] if legacy else current


def has_bibs(el: E.EntryList) -> bool:
    """Whether the export already says what everybody's dorsale is.

    All of them or none: a file that numbers half its riders is one somebody
    has been editing, and renumbering it would move the ones already printed on
    a start list.
    """
    return bool(el.riders) and all(r.bib for r in el.riders.values())


def missing_bibs(el: E.EntryList) -> list[str]:
    """Who has no dorsale, by name - what the warning is about."""
    return [r.full_name for r in el.riders.values() if not r.bib]


def numbered(el: E.EntryList, comp: Competition, how: str) -> E.EntryList:
    """Deal out the dorsali. The list is returned changed in place.

    Three ways, and the choice is the jury's because all three are ridden
    somewhere: straight down the file as it arrived, continuously but grouped
    by categoria, or from 1 again in every categoria. The last one is only
    usable where the categorie never line up together - two riders share a
    number the moment they do.
    """
    if how not in NUMBERINGS:
        return el
    riders = list(el.riders.values())
    if how == AS_IMPORTED:
        for n, rider in enumerate(riders, start=1):
            rider.bib = n
        return el

    order = comp.cat_order()
    by_cat: dict[str, list] = {}
    for rider in riders:
        by_cat.setdefault(rider.cat, []).append(rider)
    # a categoria the programme does not declare still gets numbers, after the
    # ones it does: the file is what it is, and a rider without a dorsale is
    # worse than a rider in a categoria nobody will race
    cats = [c for c in order if c in by_cat] + [c for c in by_cat
                                                if c not in order]
    n = 1
    for cat in cats:
        if how == BY_CAT_RESTART:
            n = 1
        for rider in by_cat[cat]:
            rider.bib = n
            n += 1
    return el


# ── replacing the file: what the new one changes ────────────────────────────
#
# A corrected elenco arrives at every championship - a region that entered two
# riders late, a categoria keyed wrong, a certificate that turned up. Reading it
# over the workbook must not cost the giuria the afternoon it spent ticking
# event, so the two are merged rather than one dropped on the other: the
# **file** says who is entered and what the federation knows about them, the
# **workbook** says what the giuria has decided about them.

#: What is compared between the two, and reported when it moves. Not every
#: field: a note reworded or a certificate reissued is not something a jury
#: needs told, and a delta that lists everything is one nobody reads.
WATCHED = ("cat", "bib", "last_name", "first_name", "club", "region")


@dataclass
class Delta:
    """What replacing the entry file does, before it is done.

    Held rather than printed: the page shows it and only then writes, so an
    import that turns out to be the wrong file is one nobody has to undo.
    """

    added: list = field(default_factory=list)       # riders only the file has
    removed: list = field(default_factory=list)     # riders only the book has
    changed: list = field(default_factory=list)     # (old, new, [field, ...])
    kept_marks: int = 0    # riders whose events came across
    kept_checks: int = 0   # riders whose verifica came across

    @property
    def touched(self) -> int:
        return len(self.added) + len(self.removed) + len(self.changed)


def merge(old: E.EntryList, new: E.EntryList) -> tuple[E.EntryList, Delta]:
    """The arriving list, with the giuria's work carried onto it.

    The file wins on everything it states - categoria, dorsale, società, the
    federal data - and the workbook wins on everything only the giuria knows:
    the X marks, the verifica, the NP. A dorsale is the one exception in the
    other direction: an export that numbers nobody (Fattore K sends none) must
    not wipe the numbers already printed on a start list.

    Riders are matched by UCI ID, which is what `rider_key` uses whenever there
    is one, so a rider who has changed categoria is still the same rider. One
    without an ID is matched on the key alone - `cat-bib` - and a change of
    either shows up as one entered and one gone, which is the truth as far as
    the two files can tell it.
    """
    delta = Delta()
    was = {_match_key(r): r for r in old.riders.values()}
    seen = set()

    for rider in new.riders.values():
        match = _match_key(rider)
        before = was.get(match)
        if before is None:
            delta.added.append(rider)
            continue
        seen.add(match)
        if before.events:
            rider.events = {code: e for code, e in before.events.items()}
            delta.kept_marks += 1
        if before.not_starting:
            rider.not_starting = before.not_starting
            delta.kept_checks += 1
        rider.bib = rider.bib or before.bib
        # compared *after* the carry-over, or an export that numbers nobody
        # would report every rider in the meeting as having changed dorsale
        moved = [f for f in WATCHED
                 if _s(getattr(before, f)) != _s(getattr(rider, f))]
        if moved:
            delta.changed.append((before, rider, moved))

    delta.removed = [r for k, r in was.items() if k not in seen]
    return new, delta


def _match_key(rider) -> str:
    """What makes two rows the same rider across two files."""
    return rider.uci_id or rider.key


def _s(value) -> str:
    return str(value if value is not None else "").strip().upper()


def events_of(comp: Competition, cat: str) -> list[str]:
    """The event that categoria rides, as columns of its sheet."""
    return [e for e in comp.events_for(cat) if e != EVENT_ENTRY_LIST]


def build(el: E.EntryList, comp: Competition, path: str | Path) -> Path:
    """Write the workbook this competition is run from.

    `KSPORT` first - the export, whole, plus the two columns the giuria fills
    in - and then a sheet per categoria that rides something, with a column per
    event it rides. The ticks already in the list are written into those
    columns, so building it again over a list that has been worked on does not
    lose the work.
    """
    import openpyxl

    path = Path(path)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    _write_ksport(wb, el, comp)
    for cat in comp.cat_order():
        riders = _riders_of(el, cat)
        if not riders or not events_of(comp, cat):
            continue
        _write_category(wb, comp, cat, riders)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return path


def sync(path: str | Path, comp: Competition) -> Path:
    """Write the workbook again for a programme that has moved.

    A categoria added, an event ticked or unticked: the sheets and their
    columns follow, and **everything anybody has written stays** - the file is
    read back into an entry list first (`entries.import_master`), so the X
    marks, the dorsali and the check-in come back out on the other side. A
    column whose event is no longer ridden goes; one whose event is
    new arrives empty, which is exactly the work left to do.
    """
    path = Path(path)
    el = E.import_master(path, comp)
    return build(el, comp, path)


def _riders_of(el: E.EntryList, cat: str) -> list:
    return sorted(el.by_cat(cat), key=lambda r: (r.bib or 9999, r.last_name))


def _write_ksport(wb, el: E.EntryList, comp: Competition) -> None:
    """The federal export, kept whole, with the giuria's two columns after it."""
    sheet = comp.entry_sheet
    fields = list(dict.fromkeys(sheet.ksport.values())) or list(sheet.fields)
    checks = list(dict.fromkeys(sheet.check_in.values())) or ["checked_in",
                                                              "not_starting"]
    ws = wb.create_sheet(KSPORT)
    ws.append([sheet.header_of(f, ksport=True) for f in fields]
              + [sheet.header_of(f) for f in checks])
    for rider in sorted(el.riders.values(),
                        key=lambda r: (r.bib or 9999, r.last_name)):
        ws.append([_cell(rider, f) for f in fields]
                  + [_flag(getattr(rider, f, False)) for f in checks])


def _write_category(wb, comp: Competition, cat: str, riders: list) -> None:
    """One categoria: the fixed columns, then a column per event it rides.

    The header is the **first row**. The federation's own template carries five
    empty ones above it - a letterhead - and a file this app writes has no
    reason to: `entry_formats.master` says so, and that is the layout the
    programme is left pointing at once the workbook exists.
    """
    sheet = comp.entry_sheet
    events = events_of(comp, cat)
    ws = wb.create_sheet(cat)
    ws.append([sheet.header_of(f) for f in sheet.fields]
              + [comp.event(e).short for e in events]
              + [sheet.header_of(f) for f in
                 dict.fromkeys(sheet.check_in.values())])
    for rider in riders:
        row = [_cell(rider, f) for f in sheet.fields]
        for event in events:
            entry = rider.events.get(event)
            row.append(entry.flag if entry else "")
        row += [_flag(getattr(rider, f, False))
                for f in dict.fromkeys(sheet.check_in.values())]
        ws.append(row)


def _cell(rider, name: str):
    """One value as the workbook writes it - never the dataclass default."""
    if name == "cat":
        return "NP" if rider.not_starting else rider.cat
    value = getattr(rider, name, "")
    return "" if value is None else value


def _flag(value: bool) -> str:
    """`SI`, the way the file is read back (`entries.YES_VALUES`)."""
    return E.YES if value else ""
