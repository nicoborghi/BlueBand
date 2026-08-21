"""Entry-list tests pinned to the real CITA 26 master workbook.

The event counts below were read off the workbook column by column; they
are the contract the importer must keep across refactors.
"""

import pytest

from dataclasses import replace
from pathlib import Path

from core import entries as E
from core.config import Quotas
from core.entries import (Patch, apply_overlay, build_teams_and_pairs,
                          effective_entries, export_xlsx, import_master,
                          norm_cat, parse_flag, save_import, save_overlay,
                          validate_entries, check_in_progress)
from core.models import EntryList, Rider, EventEntry


@pytest.fixture(scope="session")
def entries(iscritti_path, comp):
    return import_master(iscritti_path, comp)


# ── parsing primitives ──────────────────────────────────────────────────────

def test_parse_flag():
    assert parse_flag("X").starter and parse_flag("X").pair is None
    assert parse_flag("R").starter is False
    assert parse_flag("X RISERVA").starter is False
    assert parse_flag("1").pair == 1
    assert parse_flag(2).pair == 2
    # letters are the notation the app writes; the workbook's digits still read
    assert parse_flag("A").pair == 1 and parse_flag("A").flag == "A"
    assert parse_flag("b").pair == 2 and parse_flag("2").flag == "B"
    assert parse_flag("C").pair == 3
    assert parse_flag("") is None
    assert parse_flag(None) is None
    assert parse_flag("  ") is None  # stray spaces in AL 'Omnium '


def test_norm_cat():
    assert norm_cat("AL") == "AL"
    assert norm_cat("ALLIEVO") == "AL"
    assert norm_cat("ESORDIENTE F. ") == "ED"  # trailing space, regional form
    assert norm_cat("Allievo f.") == "DA"


# ── master workbook ─────────────────────────────────────────────────────────

def test_rider_counts(entries):
    assert len(entries.riders) == 238  # == 'LISTA UNICA ATLETI'
    per_cat = {c: len(entries.by_cat(c)) for c in ("ES", "ED", "AL", "DA")}
    assert per_cat == {"ES": 54, "ED": 34, "AL": 99, "DA": 51}


def test_every_rider_has_a_region(entries):
    assert [r.full_name for r in entries.riders.values() if not r.region] == []


@pytest.mark.parametrize("cat,event,starters,reserves", [
    ("ES", "omnium", 34, 0), ("ES", "velocita", 16, 0), ("ES", "madison", 28, 0),
    ("ED", "omnium", 25, 0), ("ED", "velocita", 9, 0), ("ED", "madison", 18, 0),
    ("AL", "vel_squadre", 21, 5), ("AL", "ins_squadre", 43, 6),
    ("AL", "omnium", 44, 0), ("AL", "velocita", 27, 0),
    ("AL", "keirin", 36, 1), ("AL", "ins_individuale", 30, 0),
    ("AL", "madison", 36, 1),
    ("DA", "vel_squadre", 15, 0), ("DA", "ins_squadre", 20, 2),
    ("DA", "omnium", 24, 0), ("DA", "velocita", 15, 0),
    ("DA", "keirin", 20, 0), ("DA", "ins_individuale", 18, 0),
    ("DA", "madison", 20, 1),
])
def test_event_matrix(entries, cat, event, starters, reserves):
    riders = [r for r in entries.by_cat(cat) if event in r.events]
    assert sum(1 for r in riders if r.events[event].starter) == starters
    assert sum(1 for r in riders if not r.events[event].starter) == reserves


def test_ksport_enrichment(entries):
    with_fci = [r for r in entries.riders.values() if r.fci_code]
    assert len(with_fci) == 227  # KSPORT sheet has 227 rows
    r = next(r for r in entries.riders.values() if r.last_name == "TASSINARI")
    assert r.cat == "ES" and r.province == "RA" and r.region == "EMILIA ROMAGNA"
    assert r.birth_date and r.fci_code


# ── derived entities ────────────────────────────────────────────────────────

def test_teams_are_never_split_automatically(entries):
    """Plain `X` is one squadra per region, however many riders carry it.

    The workbook only says who is in the event; cutting a region into A and B
    is the jury's call at the check-in, so an oversized region stays one team
    and is reported instead.
    """
    al = [t for t in entries.teams.values()
          if t.cat == "AL" and t.event == "ins_squadre" and t.riders]
    labels = {t.label for t in al}
    assert "LOMBARDIA" in labels and "LAZIO" in labels
    assert not any(name.startswith("LOMBARDIA ") for name in labels)
    big = next(t for t in al if t.label == "LOMBARDIA")
    assert len(big.riders) > 4
    assert any("LOMBARDIA" in e and "invece di 4" in e
               for e in entries.errors)


def test_team_letter_can_be_chosen_in_the_entry_flag(entries, comp):
    """`1`/`2` in the event cell pins a rider to squadra A or B."""
    riders = sorted([r for r in entries.riders.values()
                     if r.cat == "AL" and r.region == "LOMBARDIA"
                     and "ins_squadre" in r.events],
                    key=lambda r: r.bib or 0)
    keep = {r.key: r.events["ins_squadre"] for r in riders}
    try:
        for i, r in enumerate(riders):  # odd bibs to A, even ones to B
            r.events["ins_squadre"] = EventEntry(starter=True,
                                                 pair=2 if i % 2 else 1)
        build_teams_and_pairs(entries, comp)
        by_label = {t.label: [entries.riders[k].bib for k in t.riders]
                    for t in entries.teams.values()
                    if t.cat == "AL" and t.event == "ins_squadre"
                    and t.region == "LOMBARDIA"}
        assert by_label["LOMBARDIA A"] == [r.bib for r in riders[::2]]
        assert by_label["LOMBARDIA B"] == [r.bib for r in riders[1::2]]
    finally:
        for r in riders:
            r.events["ins_squadre"] = keep[r.key]
        build_teams_and_pairs(entries, comp)


def test_a_lettered_squadra_wants_exactly_four_starters(entries, comp):
    """`A` rides, `AR` is its riserva - anything else is an error."""
    riders = sorted([r for r in entries.riders.values()
                     if r.cat == "AL" and r.region == "LOMBARDIA"
                     and "ins_squadre" in r.events],
                    key=lambda r: r.bib or 0)[:6]
    keep = {r.key: r.events["ins_squadre"] for r in riders}
    try:
        for i, r in enumerate(riders[:5]):  # four A plus one AR
            r.events["ins_squadre"] = EventEntry(starter=i < 4, pair=1)
        riders[5].events["ins_squadre"] = EventEntry(starter=True, pair=2)
        build_teams_and_pairs(entries, comp)
        team = entries.teams["AL:ins_squadre:LOMBARDIA:A"]
        assert [entries.riders[k].bib for k in team.riders] == \
            [r.bib for r in riders[:4]]
        assert [entries.riders[k].bib for k in team.reserves] == [riders[4].bib]
        assert entries.riders[riders[4].key].events["ins_squadre"].flag == "AR"
        # squadra B fields one rider: that is an error, not a warning
        assert any("squadra B" in e and "1 titolari invece di 4" in e
                   for e in entries.errors)
    finally:
        for r in riders:
            r.events["ins_squadre"] = keep[r.key]
        build_teams_and_pairs(entries, comp)


# ── two rappresentative riding as one squadra (deroga) ──────────────────────

def _merged(comp, events=("ins_squadre",)):
    """The competition with the Piemonte / Valle d'Aosta deroga in force."""
    sheet = replace(comp.entry_sheet,
                    team_merge={"PIEMONTE": "PIEMONTE - V.D.A",
                                "VALLE D'AOSTA": "PIEMONTE - V.D.A"},
                    team_merge_events=list(events))
    return replace(comp, entry_sheet=sheet)


def _two_regions() -> EntryList:
    """Two riders of Piemonte and two of Valle d'Aosta, in both team events."""
    el = EntryList()
    for i, region in enumerate(("PIEMONTE", "VALLE D'AOSTA")):
        for j in range(2):
            bib = 10 * (i + 1) + j
            el.riders[str(bib)] = Rider(
                key=str(bib), bib=bib, cat="DA", region=region,
                last_name=f"R{bib}",
                events={"ins_squadre": EventEntry(starter=True),
                        "madison": EventEntry(starter=True)})
    return el


def test_two_regions_ride_as_one_squadra_where_authorised(comp):
    """The deroga composes one quartetto out of two rappresentative."""
    el = _two_regions()
    build_teams_and_pairs(el, _merged(comp))
    teams = [t for t in el.teams.values() if t.event == "ins_squadre"]
    assert [t.label for t in teams] == ["PIEMONTE - V.D.A"]
    assert len(teams[0].riders) == 4        # and so no wrong-size error
    assert not el.errors
    # the riders keep their own regione: only the squadra is joint
    assert {r.region for r in el.riders.values()} == {"PIEMONTE",
                                                      "VALLE D'AOSTA"}


def test_the_deroga_holds_only_where_it_was_granted(comp):
    """Authorised for the inseguimento: the madison still pairs by region."""
    el = _two_regions()
    build_teams_and_pairs(el, _merged(comp))
    assert {p.region for p in el.pairs.values()} == {"PIEMONTE",
                                                     "VALLE D'AOSTA"}
    # no event listed means every team event, the madison with them
    el = _two_regions()
    build_teams_and_pairs(el, _merged(comp, events=()))
    assert {p.region for p in el.pairs.values()} == {"PIEMONTE - V.D.A"}


def test_without_the_deroga_a_region_is_itself(comp):
    """No `team_merge`: two regions of two, and two undersized squadre."""
    el = _two_regions()
    plain = replace(comp, entry_sheet=replace(comp.entry_sheet, team_merge={}))
    build_teams_and_pairs(el, plain)
    assert {t.region for t in el.teams.values() if t.event == "ins_squadre"} \
        == {"PIEMONTE", "VALLE D'AOSTA"}
    assert len(el.errors) == 2


def test_madison_pairs(entries):
    pairs = [p for p in entries.pairs.values() if p.cat == "ED"]
    assert pairs
    assert all(len(p.riders) <= 2 for p in pairs)
    # pair numbers from the workbook's 1/2 column are preserved per region
    er = sorted([p for p in entries.pairs.values()
                 if p.cat == "AL" and p.region == "EMILIA ROMAGNA"],
                key=lambda p: p.number)
    assert [p.number for p in er] == [1, 2]
    # two coppie in a region are A and B, like its two quartetti
    assert [p.label for p in er] == ["EMILIA ROMAGNA A", "EMILIA ROMAGNA B"]
    # a region with one coppia is just the region
    single = next(p for p in entries.pairs.values()
                  if sum(1 for q in entries.pairs.values()
                         if (q.cat, q.region) == (p.cat, p.region)) == 1)
    assert single.label == single.region


# ── validation ──────────────────────────────────────────────────────────────

def test_validation_is_short_and_specific(entries, comp):
    issues = validate_entries(entries, comp)
    codes = {i.code for i in issues}
    # real problems in the 2026 list: riders without a bib, quota overruns
    assert "bib" in codes
    assert sum(1 for i in issues if i.code == "bib") == 10
    assert any("SICILIA" in i.message for i in issues if i.code == "quota_region")
    # a squadra still to be composed is something the jury has to fix
    assert any("LOMBARDIA" in i.message and "invece di 4" in i.message
               for i in issues if i.code == "teams" and i.level == "error")
    # the certificate column is an issue date, not an expiry: it must not
    # flag all 227 riders
    assert sum(1 for i in issues if i.code == "certificate") < 10
    assert len(issues) < 40


# ── overlay ─────────────────────────────────────────────────────────────────

def test_overlay_applies_and_reports_stale(iscritti_path, comp):
    el = import_master(iscritti_path, comp)
    key = next(r.key for r in el.by_cat("AL") if "omnium" in r.events)
    stale = apply_overlay(el, [
        Patch(target=key, op="clear_spec", field="omnium", reason="ritirato"),
        Patch(target=key, op="set_field", field="bib", value=999),
        Patch(target="INESISTENTE", op="set_np", value=True, reason="x"),
    ], comp)
    assert "omnium" not in el.riders[key].events
    assert el.riders[key].bib == 999
    assert len(stale) == 1 and "INESISTENTE" in stale[0]


def test_overlay_survives_reimport(store, iscritti_path, comp):
    el = import_master(iscritti_path, comp)
    key = next(r.key for r in el.by_cat("DA") if r.bib)
    save_import(store, el)
    save_overlay(store, [Patch(target=key, op="set_np", value=True, reason="malato")])

    # a fresh import of the same workbook must keep the jury's decision
    save_import(store, import_master(iscritti_path, comp))
    eff, stale = effective_entries(store, comp)
    assert stale == []
    assert eff.riders[key].not_starting is True
    assert len(eff.riders) == 238
    assert eff.teams and eff.pairs  # rebuilt after the overlay


def test_the_overlay_can_be_switched_off_and_the_file_rules(store, iscritti_path,
                                                            comp):
    """Impostazioni: the entry list is the workbook and nothing else.

    The way to work when the file is the master - iscritti, dorsali and
    specialità are changed there and re-imported. The patches are set aside,
    not thrown away: switched back on, the same decision is in force again.
    """
    el = import_master(iscritti_path, comp)
    key = next(r.key for r in el.by_cat("DA") if r.bib)
    save_import(store, el)
    save_overlay(store, [Patch(target=key, op="set_np", value=True,
                               reason="malato")])

    E.set_overlay_on(store, False)
    off, stale = effective_entries(store, comp)
    assert E.overlay_on(store) is False
    assert off.riders[key].not_starting is False and stale == []
    # the squadre and the coppie are read off the entries either way
    assert off.teams and off.pairs

    E.set_overlay_on(store, True)
    on, _ = effective_entries(store, comp)
    assert E.overlay_on(store) is True
    assert on.riders[key].not_starting is True
    # and it is on unless it was turned off: nothing set, nothing changes
    assert E.overlay_on(_NoSettings()) is True


class _NoSettings:
    settings: dict = {}


def test_the_edits_can_be_written_into_the_workbook_itself(store, iscritti_path,
                                                           comp, tmp_path):
    """With the overlay off the app edits the file: the cell, not a patch.

    Written through `Rider.source` - the sheet and row each rider was read
    from - and read back to prove the file itself says it now.
    """
    import shutil

    path = tmp_path / iscritti_path.name
    shutil.copy2(iscritti_path, path)
    el = import_master(path, comp)
    rider = next(r for r in el.by_cat("AL") if r.bib and r.events)
    event = next(iter(rider.events))

    written, refused = E.write_back(path, comp, el, [
        Patch(target=rider.key, op="set_field", field="bib", value=999,
              reason="dorsale rifatto"),
        Patch(target=rider.key, op="clear_event", field=event,
              reason="ritirata l'iscrizione"),
    ], store=store)
    assert (written, refused) == (2, [])

    again = import_master(path, comp)
    assert again.riders[rider.key].bib == 999
    assert event not in again.riders[rider.key].events
    # the file the app does not own is never written without a copy aside
    assert list((Path(store.root) / ".snapshots" / E.SOURCE_BACKUP).iterdir())


def test_the_licence_check_is_written_where_the_file_has_a_column(store, comp,
                                                                  iscritti_path,
                                                                  tmp_path):
    """Verificato and NP go into the columns the giuria added, in both sheets.

    The federation's layout has neither: they are declared in
    `entries.check_in` and written on the foglio di categoria *and* on the
    KSPORT sheet, so a re-import reads the same answer from either one.
    """
    import shutil

    if not E.check_in_columns(comp):
        pytest.skip("this programme declares no check-in columns")
    path = tmp_path / iscritti_path.name
    shutil.copy2(iscritti_path, path)
    el = import_master(path, comp)
    rider = next(r for r in el.riders.values() if r.ksport_source)
    assert rider.checked_in is False

    written, refused = E.write_back(path, comp, el, [
        Patch(target=rider.key, op="set_checked_in", value=True),
        Patch(target=rider.key, op="set_not_starting", value=True)], store=store)
    assert refused == [] and written == 4      # two flags, two sheets each

    again = import_master(path, comp)
    assert again.riders[rider.key].checked_in is True
    assert again.riders[rider.key].not_starting is True
    # and untickable again: the cell is cleared, not left saying SI
    E.write_back(path, comp, again, [
        Patch(target=rider.key, op="set_checked_in", value=False)], store=store)
    assert import_master(path, comp).riders[rider.key].checked_in is False


def test_the_licence_check_says_so_when_the_file_has_no_column(store, comp,
                                                               iscritti_path,
                                                               tmp_path):
    """Without the columns there is nowhere to put the tick, and it says which."""
    import shutil

    path = tmp_path / iscritti_path.name
    shutil.copy2(iscritti_path, path)
    el = import_master(path, comp)
    rider = next(iter(el.riders.values()))
    bare = replace(comp, entry_sheet=replace(comp.entry_sheet, check_in={}))
    assert E.check_in_columns(bare) == ()

    written, refused = E.write_back(path, bare, el, [
        Patch(target=rider.key, op="set_checked_in", value=True)], store=store)
    assert written == 0 and len(refused) == 1
    assert rider.full_name in refused[0]


def test_a_check_in_column_next_to_the_specialita_is_not_read_as_one(comp):
    """The giuria puts the two columns where there is room - even mid-sheet.

    They are known by name, so they neither cut the run of event columns short
    nor get reported as a specialità nobody recognises.
    """
    import openpyxl
    from core.entries import _events_by_header, _read_category_sheet

    sheet, cat = comp.entry_sheet, comp.cat_order()[0]
    events = [s for s in (comp.events_for(cat) or comp.event_order())
              if s != "entry_list"]
    heads = ([sheet.header_of(f) for f in sheet.fields]
             + [comp.event(s).short for s in events]
             + [sheet.header_of("checked_in"), sheet.header_of("not_starting")])
    ws = openpyxl.Workbook().active
    for c, h in enumerate(heads, start=1):
        ws.cell(sheet.header_row, c, h)
    row = {"bib": 7, "uci_id": "10000000001", "last_name": "ROSSI",
           "first_name": "MARIO", "cat": cat, "region": "TOSCANA"}
    for c, f in enumerate(sheet.fields, start=1):
        ws.cell(sheet.first_data_row, c, row.get(f, ""))
    for c in range(len(sheet.fields) + 1, len(heads) - 1):
        ws.cell(sheet.first_data_row, c, "X")
    ws.cell(sheet.first_data_row, len(heads) - 1, "SI")   # Verificato

    el = EntryList()
    _read_category_sheet(ws, cat, el, _events_by_header(comp), sheet)
    rider = el.riders["10000000001"]
    assert set(rider.events) == set(events)
    assert (rider.checked_in, rider.not_starting) == (True, False)
    assert el.warnings == []


def test_np_rider_is_excluded_from_entries(iscritti_path, comp):
    el = import_master(iscritti_path, comp)
    key = next(r.key for r in el.by_cat("AL") if "keirin" in r.events)
    before = len(el.entered("AL", "keirin"))
    apply_overlay(el, [Patch(target=key, op="set_np", value=True)], comp)
    assert len(el.entered("AL", "keirin")) == before - 1


# ── licence check: partenti are ticked, NP is separate ──────────────────────

def _rider(key, cat="AL", events=(), **kw):
    return Rider(key=key, cat=cat, bib=int(key), last_name=key.upper(),
                 region="TOSCANA",
                 events={s: EventEntry(starter=t) for s, t in events}, **kw)


def _list(*riders):
    return EntryList(riders={r.key: r for r in riders})


def test_verifica_progress_counts_what_is_left():
    el = _list(_rider("1", checked_in=True), _rider("2"),
               _rider("3", not_starting=True), _rider("4", cat="DA", checked_in=True))
    p = check_in_progress(el)
    assert (p.entries, p.verificati, p.missing, p.not_starting) == (3, 2, 1, 1)
    assert p.done is False
    assert check_in_progress(el, "DA").done is True  # only rider 4, ticked


def test_verificato_is_a_patch_and_survives_reimport(store, iscritti_path, comp):
    el = import_master(iscritti_path, comp)
    key = next(r.key for r in el.by_cat("ES") if r.bib)
    save_import(store, el)
    save_overlay(store, [Patch(target=key, op="set_verificato", value=True,
                               reason="verifica licenze")])
    save_import(store, import_master(iscritti_path, comp))
    eff, stale = effective_entries(store, comp)
    assert stale == []
    assert eff.riders[key].checked_in is True
    # verification says nothing about who races: NP is the flag that removes
    assert eff.riders[key].not_starting is False


def test_verificato_does_not_filter_the_entry_list(iscritti_path, comp):
    el = import_master(iscritti_path, comp)
    before = len(el.entered("AL", "keirin"))
    for r in el.riders.values():
        r.checked_in = False
    assert len(el.entered("AL", "keirin")) == before


# ── event-count limit (STP comunicato 016) ──────────────────────────────

def _event_issues(el, comp, **quota_kw):
    q = Quotas(max_events_per_rider={"AL": 4, "ES": 2}, **quota_kw)
    issues = validate_entries(el, replace(comp, quotas=q))
    return [i for i in issues if i.code == "quota_rider"]


def test_max_events_per_rider_is_reported_at_the_configured_level(comp):
    el = _list(_rider("1", events=[("omnium", True), ("keirin", True),
                                   ("velocita", True), ("ins_squadre", True),
                                   ("vel_squadre", True)]))
    (issue,) = _event_issues(el, comp)
    assert issue.level == "warn" and "5 specialità (max 4)" in issue.message
    assert "Omnium" in issue.message  # which ones, for the desk
    assert _event_issues(el, comp, max_events_level="error")[0].level == "error"
    assert _event_issues(el, comp, max_events_level="off") == []


def test_max_events_per_rider_counts_reserves_only_when_asked(comp):
    el = _list(_rider("1", cat="ES", events=[("omnium", True), ("velocita", True),
                                             ("madison", False)]))
    assert _event_issues(el, comp) == []
    over = _event_issues(el, comp, max_events_count_reserves=True)
    assert len(over) == 1 and "3 specialità (max 2)" in over[0].message


def test_max_events_per_rider_ignores_categories_without_a_limit(comp):
    el = _list(_rider("1", cat="DA", events=[("omnium", True), ("keirin", True),
                                             ("velocita", True), ("madison", True),
                                             ("ins_squadre", True)]))
    assert _event_issues(el, comp) == []


def test_cita26_programme_sets_the_stp_limits(comp):
    q = comp.quotas
    assert q.max_events_per_rider == {"ES": 2, "ED": 2, "AL": 4, "DA": 4}
    assert q.max_events_level == "error"
    assert q.max_events_count_reserves is False


# ── export ──────────────────────────────────────────────────────────────────

def test_export_xlsx_roundtrips(entries, comp, tmp_path):
    import openpyxl
    out = export_xlsx(entries, comp, tmp_path / "iscritti.xlsx")
    wb = openpyxl.load_workbook(out)
    assert wb.sheetnames == ["ES", "ED", "AL", "DA"]
    ws = wb["AL"]
    assert ws.max_row == 100  # header + 99 riders
    assert [c.value for c in ws[1]][:4] == ["Dors.", "UCI ID", "Cognome", "Nome"]


def test_more_than_two_x_leaves_the_madison_pairing_to_be_confirmed(entries, comp):
    """Four riders with a bare X: the app paired them, the jury must confirm."""
    guessed = E.guessed_pairings(entries, comp)
    assert guessed, "l'elenco 2026 ha regioni con quattro X in madison"
    assert all(len(riders) > 2 for _c, _e, _r, riders in guessed)
    assert all(comp.event(ev).fmt == "madison" for _c, ev, _r, _ in guessed)
    # in bib order, which is exactly how _build_pairs cut them into coppie
    for _c, _e, _r, riders in guessed:
        assert [r.bib for r in riders] == sorted(r.bib for r in riders)

    warned = {(i.message.split("]")[0].strip("["), i.message.split("]")[1].split(":")[0].strip())
              for i in E.validate_entries(entries, comp) if i.code == "pairs"}
    for cat, event, region, _riders in guessed:
        assert (f"{cat} {comp.event(event).short}", region) in warned


def test_a_declared_coppia_is_not_a_guess(entries, comp):
    """`1`/`2` in the madison cell says who rides with whom: nothing to confirm."""
    cat, event, region, riders = E.guessed_pairings(entries, comp)[0]
    keep = {r.key: r.events[event] for r in riders}
    try:
        for i, r in enumerate(riders):
            r.events[event] = EventEntry(starter=True, pair=(i // 2) + 1)
        build_teams_and_pairs(entries, comp)
        left = [g for g in E.guessed_pairings(entries, comp)
                if (g[0], g[2]) == (cat, region)]
        assert not left
    finally:
        for r in riders:
            r.events[event] = keep[r.key]
        build_teams_and_pairs(entries, comp)


# ── the flat federal export ─────────────────────────────────────────────────
#
# `Iscritti_NNNNNN_KSPORT.xlsx` is what the federation actually sends: one
# sheet, one row per rider, a Categoria column - and no specialità at all.
# Which specialità a categoria runs is the programme's business, and who is
# entered in them is the jury's, ticked at the verifica and kept in the
# overlay. Re-importing therefore has to leave all of that standing.

def test_the_flat_export_is_read_without_any_category_sheet(ksport_path, comp):
    el = E.import_entries(ksport_path, comp)

    assert E.is_flat_export(ksport_path, comp) is True
    assert len(el.riders) == 257
    # keyed by UCI ID, which is the code that never changes
    assert all(r.key == r.uci_id for r in el.riders.values() if r.uci_id)
    # the categoria is read off its own column, not off the sheet a rider is on
    assert {r.cat for r in el.riders.values()} >= set(comp.cat_order())
    # no event columns in this file: nobody is entered in anything yet
    assert all(not r.events for r in el.riders.values())

    r = el.riders["10089692648"]
    assert (r.last_name, r.first_name, r.cat, r.bib) == ("MATTOSCIO", "FLAVIO",
                                                         "ES", 1)
    assert (r.club_code, r.province) == ("07L1885", "RN")
    assert r.region == "EMILIA ROMAGNA"        # the column, not the short note
    assert r.birth_date == "2012-05-22"        # a date, not a timestamp


def test_the_flat_export_says_what_it_could_not_place(ksport_path, comp):
    """Two riders are entered under a categoria the championship does not run."""
    el = E.import_entries(ksport_path, comp)
    unknown = [w for w in el.warnings if "categoria" in w]
    assert len(unknown) == 2 and all("REG" in w for w in unknown)

    # and the riders whose regione the federal file leaves as "?" are an error,
    # not a rappresentativa called "?" that pools them together
    missing = [i for i in E.validate_entries(el, comp) if i.code == "region"]
    assert missing and all("Regione" in i.message for i in missing)
    assert not any(r.region == "?" for r in el.riders.values())


def test_reimporting_the_flat_export_keeps_what_the_jury_typed(store,
                                                               ksport_path,
                                                               comp):
    """The whole point of the overlay: a reload must break nothing.

    Dorsale, regione and specialità are all things the jury fixes in the app on
    top of this file. They are patches against the UCI ID, so a second import -
    of a newer export, or of the same one - re-applies every one of them.
    """
    el = E.import_entries(ksport_path, comp)
    key = next(r.key for r in el.by_cat("AL") if r.uci_id)
    save_import(store, el)
    save_overlay(store, [
        Patch(target=key, op="set_event", field="keirin", value="X",
              reason="iscritto in verifica"),
        Patch(target=key, op="set_field", field="bib", value=201,
              reason="dorsale corretto"),
        Patch(target=key, op="set_field", field="region", value="CAMPANIA",
              reason="regione mancante nel file"),
        Patch(target=key, op="set_checked_in", value=True),
    ])

    save_import(store, E.import_entries(ksport_path, comp))
    eff, stale = effective_entries(store, comp)
    assert stale == []
    rider = eff.riders[key]
    assert rider.events["keirin"].starter is True
    assert (rider.bib, rider.region, rider.checked_in) == (201, "CAMPANIA", True)
    assert eff.entered("AL", "keirin") == [rider]


def test_the_importer_picks_the_reader_the_file_asks_for(iscritti_path,
                                                         ksport_path, comp):
    """One button, two shapes: the master still has its category sheets."""
    assert E.is_flat_export(iscritti_path, comp) is False
    master = E.import_entries(iscritti_path, comp)
    assert len(master.riders) == 238
    assert any(r.events for r in master.riders.values())


def test_the_missing_squadra_is_the_one_this_competition_groups_by(comp):
    """At an open meeting a rider without a società is the incomplete one."""
    by_club = replace(comp, entry_sheet=replace(comp.entry_sheet,
                                                team_group="club"))
    el = _list(_rider("1", club=""), _rider("2", club="GS Pippo"))
    missing = [i for i in validate_entries(el, by_club) if i.code == "region"]
    assert len(missing) == 1 and "Società" in missing[0].message

    # the same list at a championship: both have their regione, nothing to say
    assert not [i for i in validate_entries(el, comp) if i.code == "region"]


def test_an_entry_the_programme_does_not_run_is_reported(comp):
    """Typed by hand at the verifica: a tick in the wrong column is a finding."""
    el = _list(_rider("1", cat="ES", events=[("keirin", True)]))
    assert "keirin" not in comp.events_for("ES")
    found = [i for i in validate_entries(el, comp) if i.code == "event_not_run"]
    assert len(found) == 1 and "Keirin" in found[0].message
