"""Building the elenco iscritti of a competition from what the federation sends.

The federal export is a flat list with no specialità in it; what a meeting is
run from is a workbook with a sheet per categoria and a column per specialità
of that categoria. These are about that translation - and about the two things
it must never get wrong: the dorsali, and the work already done in the file.
"""

from pathlib import Path

import pytest

from conftest import EXAMPLE_ENTRIES, EXAMPLE_PROGRAMME
from core import entries as E
from core import entry_book as B
from core import entry_formats as F
from core import programme as P
from core.config import load_competition


@pytest.fixture
def comp():
    """A competition whose programme says nothing about its own file layout.

    Which is every competition being set up: the mapping used to have to be
    written into `entries:` by hand before anything could be imported at all,
    and now it comes from the format (`core.entry_formats`).
    """
    return F.applied(load_competition(EXAMPLE_PROGRAMME), "ksport")


@pytest.fixture
def master():
    """The same competition, read as the workbook *this app writes*.

    Two layouts and two jobs: the arriving file is a ksport export with its
    header five rows down, and what is built from it is ours - a sheet per
    categoria, the header on the first row.
    """
    return F.applied(load_competition(EXAMPLE_PROGRAMME), "master")


@pytest.fixture
def entries(comp):
    """The fictional field that ships with the repo (`example`).

    Which is the point of it being fictional: a real elenco is a few hundred
    minors' personal data and cannot be in the repo, so these used to run on
    the one laptop that has the Drive folder mounted and skip everywhere else.
    """
    return E.import_ksport_export(EXAMPLE_ENTRIES, comp)


# ── the format, which is what makes the file readable at all ────────────────

def test_a_competition_that_states_no_layout_is_read_by_the_format(comp):
    """Nothing had to be typed into `programme.yaml` for this to be readable."""
    bare = load_competition(EXAMPLE_PROGRAMME)
    assert not bare.entry_sheet.ksport, "the fixture is not testing anything"
    assert comp.entry_sheet.ksport["CodiceUci"] == "uci_id"
    assert comp.entry_sheet.check_in == {"Verificato": "checked_in",
                                         "NP": "not_starting"}


def test_what_a_programme_states_itself_wins_over_the_format(tmp_path):
    """A federation that renames a column is a line in the table; a meeting
    that reads a file of its own is a line in its programme.

    And it wins **whole**. A mapping describes one file: keeping the table's
    answer for the fields this one does not mention would leave two headers
    pointing at the same field, and the import would take whichever of the two
    columns came first in the file.
    """
    written = tmp_path / "programme.yaml"
    written.write_text("name: Una riunione che legge un file suo\n"
                       "entries:\n"
                       "  ksport:\n"
                       '    "Tessera UCI": uci_id\n', encoding="utf-8")
    comp = load_competition(written)
    mine = dict(comp.entry_sheet.ksport)
    assert mine == {"Tessera UCI": "uci_id"}
    assert F.applied(comp, "ksport").entry_sheet.ksport == mine
    # and a competition that states nothing still gets the table's, whole
    bare = load_competition(EXAMPLE_PROGRAMME)
    assert F.applied(bare, "ksport").entry_sheet.ksport \
        == F.layout("ksport")["ksport"]


def test_the_sample_export_is_read_whole(entries):
    assert len(entries.riders) == 140
    assert not entries.warnings, "the example file must read without a complaint"
    assert {r.cat for r in entries.riders.values()} == {"ES", "ED", "AL", "DA",
                                                        "JU", "DJ"}


# ── the dorsali ─────────────────────────────────────────────────────────────

def test_a_file_that_numbers_nobody_is_not_a_file_with_numbers(entries):
    """All of them or none: a half-numbered file is one somebody has edited."""
    assert not B.has_bibs(entries)
    assert len(B.missing_bibs(entries)) == 1


def test_the_bibs_can_be_dealt_out_down_the_file(entries, comp):
    B.numbered(entries, comp, B.AS_IMPORTED)
    assert B.has_bibs(entries)
    assert sorted(r.bib for r in entries.riders.values()) == list(
        range(1, len(entries.riders) + 1))


def test_the_bibs_can_run_on_from_one_categoria_to_the_next(entries, comp):
    """1…N, N+1…M: one run, in the order the programme lists the categorie."""
    B.numbered(entries, comp, B.BY_CAT_RUNNING)
    assert B.has_bibs(entries)
    numbers = sorted(r.bib for r in entries.riders.values())
    assert numbers == list(range(1, len(numbers) + 1))
    first = [c for c in comp.cat_order()
             if any(r.cat == c for r in entries.riders.values())][0]
    assert min(r.bib for r in entries.riders.values() if r.cat == first) == 1
    # and no categoria interleaves with another
    for cat in comp.cat_order():
        bibs = sorted(r.bib for r in entries.riders.values() if r.cat == cat)
        if bibs:
            assert bibs == list(range(bibs[0], bibs[0] + len(bibs)))


def test_the_bibs_can_start_again_in_every_categoria(entries, comp):
    """Which is only usable where two categorie never line up together."""
    B.numbered(entries, comp, B.BY_CAT_RESTART)
    for cat in comp.cat_order():
        bibs = sorted(r.bib for r in entries.riders.values() if r.cat == cat)
        if bibs:
            assert bibs == list(range(1, len(bibs) + 1))


# ── the workbook ────────────────────────────────────────────────────────────

def test_the_workbook_has_the_federal_sheet_and_one_per_categoria(
        entries, comp, master, tmp_path):
    import openpyxl

    B.numbered(entries, comp, B.BY_CAT_RUNNING)
    out = B.build(entries, master, tmp_path / "Iscritti.xlsx")
    wb = openpyxl.load_workbook(out)
    assert wb.sheetnames[0] == B.KSPORT
    riding = [c for c in comp.cat_order() if B.events_of(comp, c)
              and any(r.cat == c for r in entries.riders.values())]
    assert wb.sheetnames[1:] == riding

    # one column per specialità that categoria rides, and the two the giuria
    # fills in - which the federation's own file has no place for. The header
    # is the first row: the empty ones above it in the federation's template
    # are a letterhead, and this file is ours.
    assert master.entry_sheet.header_row == 1
    ws = wb[riding[0]]
    headers = [c.value for c in ws[master.entry_sheet.header_row]]
    events = [comp.event(e).short for e in B.events_of(comp, riding[0])]
    assert headers[-len(events) - 2:] == events + ["Verificato", "NP"]


def test_the_workbook_is_read_back_by_the_importer(entries, comp, master,
                                                   tmp_path):
    """It is written to be read: the same importer, and nothing lost."""
    B.numbered(entries, comp, B.BY_CAT_RUNNING)
    out = B.build(entries, master, tmp_path / "Iscritti.xlsx")
    back = E.import_master(out, master)
    assert len(back.riders) == len(entries.riders)
    assert B.has_bibs(back)
    assert {r.full_name for r in back.riders.values()} == {
        r.full_name for r in entries.riders.values()}


def test_following_the_programme_keeps_the_work_already_done(
        entries, comp, master, tmp_path):
    """A specialità added is a column added, and nothing anybody ticked moves.

    Which is the whole reason the file is read back before it is written
    again: a rebuild that started from the export would throw away the
    iscrizioni, the dorsali and the verifica of everybody in it.
    """
    B.numbered(entries, comp, B.BY_CAT_RUNNING)
    out = B.build(entries, master, tmp_path / "Iscritti.xlsx")

    worked = E.import_master(out, master)
    rider = next(r for r in worked.riders.values() if r.cat == "ES")
    rider.events["omnium"] = E.EventEntry(starter=True)
    B.build(worked, master, out)

    P.add_item(master, "ES", "eliminazione", 1)
    B.sync(out, master)

    import openpyxl
    ws = openpyxl.load_workbook(out)["ES"]
    headers = [c.value for c in ws[master.entry_sheet.header_row]]
    assert "Eliminazione" in headers
    back = E.import_master(out, master)
    kept = back.riders[rider.key]
    # entering the specialità is what verifies the rider: nothing to tick
    assert kept.events["omnium"].starter and kept.checked_in
    assert kept.bib == rider.bib


# ── replacing the file ──────────────────────────────────────────────────────
#
# A corrected elenco arrives at every championship. What it must never cost is
# the afternoon the giuria spent ticking specialità in the workbook.

@pytest.fixture
def built(entries, master, tmp_path):
    """A workbook with a rider the giuria has already worked on.

    The dorsali are the ones the export carries, not dealt out again: what is
    under test here is the merge, and a renumbering would make every rider in
    the meeting differ from the file for a reason that is not the file's.
    """
    out = B.build(entries, master, tmp_path / B.FILENAME)
    worked = E.import_master(out, master)
    rider = next(r for r in worked.riders.values() if r.cat == "ES")
    rider.events["omnium"] = E.EventEntry(starter=True)
    B.build(worked, master, out)
    return out, rider.key


def test_a_corrected_file_does_not_undo_the_ticks(built, comp, master):
    out, key = built
    fresh = E.import_ksport_export(EXAMPLE_ENTRIES, comp)
    merged, delta = B.merge(E.import_master(out, master), fresh)

    kept = merged.riders[key]
    # entering the specialità is what verifies the rider: nothing to tick
    assert kept.events["omnium"].starter and kept.checked_in
    # the specialità came across; NP is the only flag left to carry, and
    # this rider has none
    assert delta.kept_marks == 1 and delta.kept_checks == 0


def test_an_export_that_numbers_nobody_keeps_the_numbers_already_dealt(
        entries, comp, master, tmp_path):
    """Fattore K sends no dorsali: a re-import must not blank the start lists."""
    B.numbered(entries, comp, B.BY_CAT_RUNNING)
    out = B.build(entries, master, tmp_path / B.FILENAME)

    fresh = E.import_ksport_export(EXAMPLE_ENTRIES, comp)
    for rider in fresh.riders.values():
        rider.bib = None
    merged, delta = B.merge(E.import_master(out, master), fresh)
    assert B.has_bibs(merged)
    assert not delta.changed, "a dorsale carried over is not a dorsale changed"


def test_the_delta_says_who_arrived_who_left_and_what_moved(built, comp,
                                                            master):
    out, key = built
    fresh = E.import_ksport_export(EXAMPLE_ENTRIES, comp)
    gone = next(k for k in fresh.riders if k != key)
    withdrawn = fresh.riders.pop(gone)
    fresh.riders[key].club = "A.S.D. QUALCUN ALTRO"

    _, delta = B.merge(E.import_master(out, master), fresh)
    assert [r.key for r in delta.removed] == [withdrawn.key]
    assert not delta.added
    assert [(r.key, fields) for _, r, fields in delta.changed] \
        == [(key, ["club"])]


def test_a_rider_who_changes_categoria_is_still_the_same_rider(built, comp,
                                                              master):
    """Matched on the UCI ID, which is what the key is made of when there is one.

    A categoria keyed wrong at the regional office is the commonest correction
    of them all, and one that arrived as a withdrawal plus an entry would lose
    everything already ticked for that rider.
    """
    out, key = built
    fresh = E.import_ksport_export(EXAMPLE_ENTRIES, comp)
    fresh.riders[key].cat = "AL"

    merged, delta = B.merge(E.import_master(out, master), fresh)
    assert not delta.removed and not delta.added
    assert merged.riders[key].cat == "AL"
    assert merged.riders[key].events["omnium"].starter


def test_an_identical_file_changes_nothing(built, comp, master):
    out, _ = built
    merged, delta = B.merge(E.import_master(out, master),
                            E.import_master(out, master))
    assert delta.touched == 0
    assert len(merged.riders) == len(E.import_master(out, master).riders)


# ── where the workbook lives, and what its archive sheet is called ──────────

def test_the_workbook_has_one_name(tmp_path):
    assert B.book_path(tmp_path).name == "entry_list.xlsx"


def test_a_folder_set_up_before_the_rename_opens_on_the_file_it_has(tmp_path):
    """A jury mid-championship must not be told there is no elenco."""
    (tmp_path / f"{B.PREFIX}182447.xlsx").write_bytes(b"")
    assert B.book_path(tmp_path).name == f"{B.PREFIX}182447.xlsx"
    (tmp_path / B.FILENAME).write_bytes(b"")
    assert B.book_path(tmp_path).name == B.FILENAME


def test_the_archive_sheet_is_read_under_either_name(entries, comp, master,
                                                     tmp_path):
    """`KSPORT` became `_KSPORT`; a workbook written before that still reads."""
    import openpyxl

    out = B.build(entries, master, tmp_path / B.FILENAME)
    assert B.KSPORT == "_KSPORT"
    wb = openpyxl.load_workbook(out)
    wb[B.KSPORT].title = "KSPORT"
    wb.save(out)

    back = E.import_master(out, master)
    rider = next(r for r in back.riders.values() if r.uci_id)
    assert rider.birth_date, "the federal data was not read back"
    assert rider.ksport_source.startswith("KSPORT!")


# ── the flat federal export ─────────────────────────────────────────────────
#
# Fattore K and ksport are the same system under two names, and the export is
# one shape: one row per rider, the same headings. It used to be two entries in
# the table and the jury had to pick between them - two settings that did the
# same thing, which is a question with no right answer.

def test_the_federal_export_is_one_format(comp):
    """One entry in the table, and it is the one an import opens on."""
    assert F.codes() == ["ksport", "master"]
    assert F.default() == "ksport"
    assert F.is_flat("ksport")


def test_the_header_row_is_found_and_not_declared(comp, tmp_path):
    """The same export arrives bare and under a letterhead of empty rows.

    A row number in the table could only ever be right for one of the two, and
    the wrong one reads a file of nothing - which is what a `header_row: 6` did
    to a file whose headings are on the first line.
    """
    import openpyxl

    assert "header_row" not in F.layout("ksport"), "declared again"
    plain = E.import_ksport_export(EXAMPLE_ENTRIES, comp)

    src = openpyxl.load_workbook(EXAMPLE_ENTRIES).active
    wb = openpyxl.Workbook()
    ws = wb.active
    for _ in range(5):
        ws.append(["FEDERAZIONE CICLISTICA ITALIANA"])
    for row in src.iter_rows(values_only=True):
        ws.append(list(row))
    headed = tmp_path / "Iscritti_182447_KSPORT.xlsx"
    wb.save(headed)

    assert E._header_row(headed, comp) == 5
    assert len(E.import_ksport_export(headed, comp).riders) == len(plain.riders)


def test_the_squadra_falls_back_to_the_note_when_there_is_no_column():
    """"Iscrizione CR. LOMBARDIA" is where an export without `Regione` says it."""
    import dataclasses

    comp = F.applied(load_competition(EXAMPLE_PROGRAMME), "ksport")
    sheet = comp.entry_sheet
    without = dataclasses.replace(comp, entry_sheet=dataclasses.replace(
        sheet, ksport={k: v for k, v in sheet.ksport.items() if k != "Regione"}))

    el = E.import_ksport_export(EXAMPLE_ENTRIES, without)
    assert el.riders
    assert all(r.region for r in el.riders.values())


def test_the_real_federal_export_is_read(fattorek_path, comp):
    """The file the app was given to read, read - `.xls` and all."""
    el = E.import_ksport_export(fattorek_path, comp)
    assert len(el.riders) > 100
    rider = next(r for r in el.riders.values() if r.uci_id)
    assert rider.last_name and rider.club and rider.birth_date


# ── mapping the columns by hand ─────────────────────────────────────────────
#
# No table can know, in advance, which column of an arriving file is the
# squadra: an export with no `Regione` says the regione inside `Note`. So the
# jury maps it, once, and the mapping goes into the programme.

def test_a_field_can_be_read_from_any_column_of_the_file():
    """`Regione` ← `Note`: the case the dialog exists for."""
    import dataclasses

    comp = F.applied(load_competition(EXAMPLE_PROGRAMME), "ksport")
    mine = {h: f for h, f in comp.entry_sheet.ksport.items() if f != "region"}
    mine["Note"] = "region"
    mapped = dataclasses.replace(comp, entry_sheet=dataclasses.replace(
        comp.entry_sheet, ksport=mine))

    el = E.import_ksport_export(EXAMPLE_ENTRIES, mapped)
    assert el.riders
    # and read as a regione, not verbatim: the column says "Iscrizione CR.
    # LOMBARDIA" and the start lists must not
    assert all(r.region and not r.region.startswith("ISCRIZIONE")
               for r in el.riders.values())
    # the same regioni the column of its own gives, read the other way round
    plain = E.import_ksport_export(EXAMPLE_ENTRIES, comp)
    assert {r.region for r in el.riders.values()} \
        == {r.region for r in plain.riders.values()}


def test_the_mapping_is_taken_whole_and_the_old_column_stops_being_read():
    """Mapping a field elsewhere has to *unmap* it, or two columns claim it.

    The table says `Regione -> region`; a competition that says `Note ->
    region` and inherited the table's line beside it would have two headers
    for one field, and the import would take whichever came first in the file.
    """
    import dataclasses

    from core.config import load_competition as _load

    comp = _load(EXAMPLE_PROGRAMME)
    mine = {"Note": "region", "Cognome": "last_name", "CodiceUci": "uci_id",
            "Categoria": "cat", "DorsaleNumero": "bib"}
    stated = dataclasses.replace(comp, entry_sheet=dataclasses.replace(
        comp.entry_sheet, ksport=mine))
    assert F.applied(stated, "ksport").entry_sheet.ksport == mine


def test_the_columns_offered_are_the_ones_the_file_has():
    """The dialog asks the file, not the table."""
    comp = F.applied(load_competition(EXAMPLE_PROGRAMME), "ksport")
    columns = E.flat_columns(EXAMPLE_ENTRIES, comp)
    assert "Note" in columns and "CodiceUci" in columns
    assert all(not c.startswith("Unnamed:") for c in columns)


def test_the_mapping_dialog_opens_inside_the_import_expander():
    """It is reached from a button inside one, and must not blow the page up."""
    from streamlit.testing.v1 import AppTest

    def app():
        import streamlit as st

        from conftest import EXAMPLE_ENTRIES as ENTRIES
        from conftest import EXAMPLE_PROGRAMME as PROG
        from core import entry_formats as EFMT
        from core.config import load_competition as load
        from ui.pages import programme as PP

        draft = load(PROG)
        st.session_state[PP.DRAFT] = draft
        st.session_state[PP.DRAFT_OF] = "EX"
        st.session_state[PP._UPLOADED] = str(ENTRIES)
        with st.expander("Importa"):
            if st.button("apri", key="open"):
                PP._mapping_dialog(draft, EFMT.applied(draft, "ksport"))

    at = AppTest.from_function(app, default_timeout=60).run()
    at = at.button(key="open").click().run()
    assert not at.exception, at.exception
    # one row per field we can read off a flat export, seeded from the mapping
    assert len([s for s in at.selectbox if s.key.startswith("prog_map_")]) \
        == len(E.FLAT_FIELDS)
    assert at.selectbox(key="prog_map_region").value == "Regione"
    assert at.selectbox(key="prog_map_bib").value == "DorsaleNumero"


@pytest.fixture
def no_region(tmp_path):
    """The sample export with its `Regione` column taken out.

    Which is what an ordinary Fattore K export looks like: it names the regione
    nowhere, and at a meeting scored by regione that is a real problem to
    solve - once, on the Programma page, and not again under every import.

    `Note` goes with it: the "Iscrizione CR. LOMBARDIA" line is *also* the
    regione, and a file that still carries it is not a file with no regione in
    it - `test_a_regione_read_off_the_note_is_not_reported_missing` is the
    other half of this pair.
    """
    import openpyxl

    src = openpyxl.load_workbook(EXAMPLE_ENTRIES).active
    rows = [list(r) for r in src.iter_rows(values_only=True)]
    drop = rows[0].index("Regione")
    note = rows[0].index("Note")
    wb = openpyxl.Workbook()
    ws = wb.active
    for i, row in enumerate(rows):
        if i:
            row[note] = None
        ws.append([v for j, v in enumerate(row) if j != drop])
    out = tmp_path / "Iscritti_000000.xlsx"
    wb.save(out)
    return out


def test_a_regione_read_off_the_note_is_not_reported_missing(tmp_path):
    """The datum is there; which column it came out of is our business.

    The federal export names the regione in `Note` ("Iscrizione CR. LOMBARDIA")
    when it has no `Regione` column of its own, and the import reads it from
    there. Saying «colonna Regione assente: il dato non e stato importato» over
    a list where every atleta has one is telling the giuria something untrue.
    """
    import openpyxl

    src = openpyxl.load_workbook(EXAMPLE_ENTRIES).active
    rows = [list(r) for r in src.iter_rows(values_only=True)]
    drop = rows[0].index("Regione")
    wb = openpyxl.Workbook()
    ws = wb.active
    for row in rows:
        ws.append([v for i, v in enumerate(row) if i != drop])
    path = tmp_path / "Iscritti_000001.xlsx"
    wb.save(path)

    comp = F.applied(load_competition(EXAMPLE_PROGRAMME), "ksport")
    assert not comp.entry_sheet.mapped        # nobody has mapped anything
    el = E.import_ksport_export(path, comp)
    assert all(r.region for r in el.riders.values())
    assert not [w for w in el.warnings if "Regione" in w]
    assert not [i for i in E.validate_entries(el, comp) if "Regione" in i.message]


def test_a_mapped_file_stops_being_told_which_column_it_lacks(no_region):
    """The giuria answered the question; Verifica must stop asking it.

    An export with no `Regione` at a meeting scored by regione is a real thing
    - the squadra is then the società or the provincia, and that is a decision
    taken on the Programma page. Repeating «colonna Regione assente» under
    every import after the mapping has been set is telling the jury something
    it has already dealt with.
    """
    import dataclasses

    bare = load_competition(EXAMPLE_PROGRAMME)
    table = F.applied(bare, "ksport")
    assert not table.entry_sheet.mapped

    mine = {"DorsaleNumero": "bib", "Cognome": "last_name", "Nome": "first_name",
            "Categoria": "cat", "CodiceUci": "uci_id", "NomeSocieta": "club"}
    stated = F.applied(dataclasses.replace(bare, entry_sheet=dataclasses.replace(
        bare.entry_sheet, ksport=mine)), "ksport")
    assert stated.entry_sheet.mapped

    def squadra(comp):
        el = E.import_ksport_export(no_region, comp)
        return [w for w in el.warnings if "Regione" in w]

    assert squadra(table), "the fixture is not testing anything"
    assert not squadra(stated)


def test_a_column_the_import_cannot_work_without_is_still_reported():
    """Mapping is not a way of switching the checks off."""
    import dataclasses

    bare = load_competition(EXAMPLE_PROGRAMME)
    without = F.applied(dataclasses.replace(bare, entry_sheet=dataclasses.replace(
        bare.entry_sheet, ksport={"Cognome": "last_name"})), "ksport")

    el = E.import_ksport_export(EXAMPLE_ENTRIES, without)
    said = " ".join(el.warnings)
    for missing in ("UCI ID", "Cat.", "Dors."):
        assert missing in said, missing
    # and a field nothing maps to is not quoted as a column of the file
    assert "'uci_id'" not in said


def test_the_mapping_is_not_written_into_the_programme_as_a_flag():
    """`mapped` says where the mapping came from; it is not part of one."""
    import tempfile

    comp = F.applied(load_competition(EXAMPLE_PROGRAMME), "ksport")
    assert comp.entry_sheet.mapped is False
    written = Path(tempfile.mkdtemp()) / "programme.yaml"
    written.write_text(P.dump(comp), encoding="utf-8")
    assert "mapped" not in written.read_text(encoding="utf-8")
